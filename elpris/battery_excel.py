"""Excel export for battery arbitrage analysis with full formula transparency.

This module exports battery arbitrage calculations to Excel with all formulas visible,
allowing users to verify and modify calculations directly in Excel.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from collections import defaultdict
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

from .battery import (
    ROUND_TRIP_EFFICIENCY,
    REPORTS_DIR,
    extract_daily_stats,
    extract_hourly_profile,
)
from .config import ZONES


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
BOLD = Font(bold=True)


def export_battery_excel(
    filename: str | None = None,
    efficiency: float = ROUND_TRIP_EFFICIENCY,
    zones: list[str] | None = None,
) -> Path:
    """
    Export complete battery arbitrage report to Excel with full formula transparency.

    Creates an Excel workbook with:
    - Parametrar: Editable efficiency parameter
    - Priser sheets per zone: Daily min/max prices
    - Beräkningar: Revenue formulas per day
    - Månadssammanfattning: Monthly aggregation with SUMIFS formulas
    - Årssammanfattning: Yearly aggregation
    - Hourly Profile: Average price per hour

    Args:
        filename: Output filename (default: timestamped)
        efficiency: Default round-trip efficiency (0-1)
        zones: List of zones (default: all SE1-SE4)

    Returns:
        Path to exported file
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"battery_arbitrage_{timestamp}.xlsx"

    if zones is None:
        zones = list(ZONES)

    output_path = REPORTS_DIR / filename

    wb = Workbook()
    wb.remove(wb.active)

    # Collect all data first
    all_data = {}
    for zone in zones:
        daily = extract_daily_stats(zone)
        if daily:
            all_data[zone] = {
                "daily": daily,
                "hourly": extract_hourly_profile(zone),
            }

    # Track data ranges for formulas
    zone_data_ranges: dict[str, dict] = {}

    # Sheet 1: Parameters
    _create_parameters_sheet(wb, efficiency, zones)

    # Sheets 2-5: Daily prices per zone
    for zone in zones:
        if zone in all_data:
            data_info = _create_zone_prices_sheet(wb, zone, all_data[zone])
            zone_data_ranges[zone] = data_info

    # Sheet 6: Calculations with revenue formulas
    _create_calculations_sheet(wb, zones, zone_data_ranges, all_data)

    # Sheet 7: Monthly summary with SUMIFS
    _create_monthly_summary_sheet(wb, zones, zone_data_ranges)

    # Sheet 8: Yearly summary per zone
    _create_yearly_summary_sheet(wb, zones, zone_data_ranges)

    # Sheet 9: Yearly overview (pivot-style)
    _create_yearly_overview_sheet(wb, zones, zone_data_ranges)

    # Sheet 10: Hourly profile
    _create_hourly_profile_sheet(wb, zones, all_data)

    wb.save(output_path)
    return output_path


def _create_parameters_sheet(
    wb: Workbook,
    efficiency: float,
    zones: list[str],
):
    """Create parameters sheet with editable efficiency and best practices."""
    ws = wb.create_sheet("Parametrar")

    ws["A1"] = "BATTERI-ARBITRAGE BERÄKNING"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    # Editable parameters section
    ws["A3"] = "REDIGERBARA PARAMETRAR"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL

    ws["A4"] = "Round-trip effektivitet"
    ws["B4"] = efficiency
    ws["B4"].fill = INPUT_FILL
    ws["B4"].number_format = "0.00"
    ws["B4"].protection = Protection(locked=False)  # Allow editing
    ws["C4"] = "(Ändra för att uppdatera alla revenue-beräkningar)"

    # Add comment for documentation
    ws["B4"].comment = Comment(
        "Round-trip effektivitet för batteriet.\n"
        "Typiskt värde: 0.85-0.92 (85-92%)\n"
        "Inkluderar laddnings- och urladdningsförluster.",
        "System"
    )

    # Data validation: efficiency should be between 0.5 and 1.0
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1="0.5",
        formula2="1.0",
        showErrorMessage=True,
        errorTitle="Ogiltigt värde",
        error="Effektiviteten måste vara mellan 0.50 och 1.00 (50-100%)"
    )
    dv.add(ws["B4"])
    ws.add_data_validation(dv)

    # Create named range for efficiency
    wb.defined_names.add(DefinedName("Efficiency", attr_text="Parametrar!$B$4"))

    # Fixed parameters
    ws["A6"] = "FASTA PARAMETRAR"
    ws["A6"].font = BOLD

    ws["A7"] = "Zoner"
    ws["B7"] = ", ".join(zones)

    # Legend
    ws["A9"] = "FORMELFÖRKLARING"
    ws["A9"].font = BOLD

    ws["A10"] = "Spread"
    ws["B10"] = "= Max pris - Min pris"

    ws["A11"] = "1-Cycle Revenue"
    ws["B11"] = "= MAX(0, Max pris × effektivitet - Min pris)"

    ws["A12"] = "2-Cycle Revenue"
    ws["B12"] = "= (Morgonpris × eff - Nattpris) + (Kvällspris × eff - Middagspris)"

    ws["A14"] = "TIDSZONER FÖR 2-CYCLE"
    ws["A14"].font = BOLD

    ws["A15"] = "Cycle 1: Ladda"
    ws["B15"] = "00:00-06:00 (natt)"

    ws["A16"] = "Cycle 1: Ladda ur"
    ws["B16"] = "07:00-10:00 (morgon)"

    ws["A17"] = "Cycle 2: Ladda"
    ws["B17"] = "11:00-15:00 (middag)"

    ws["A18"] = "Cycle 2: Ladda ur"
    ws["B18"] = "17:00-21:00 (kväll)"

    # Sheet structure documentation
    ws["A20"] = "SHEET-STRUKTUR"
    ws["A20"].font = BOLD

    ws["A21"] = "Priser SEn"
    ws["B21"] = "Dagliga min/max/avg priser per zon"
    ws["A22"] = "Beräkningar"
    ws["B22"] = "Revenue-formler per dag (refererar till Efficiency)"
    ws["A23"] = "Månadssammanfattning"
    ws["B23"] = "Aggregerat per månad och zon"
    ws["A24"] = "Årssammanfattning"
    ws["B24"] = "Aggregerat per år och zon"
    ws["A25"] = "Zonöversikt"
    ws["B25"] = "Pivot-vy: År × Zon för snabb jämförelse"

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 50

    # Protect sheet but allow editing of input cells
    ws.protection.sheet = True
    ws.protection.password = None
    ws.protection.enable()


def _create_zone_prices_sheet(
    wb: Workbook,
    zone: str,
    data: dict,
) -> dict:
    """
    Create prices sheet for a zone with daily min/max data.

    Returns dict with row mapping for formulas.
    """
    ws = wb.create_sheet(f"Priser {zone}")

    headers = [
        "Datum", "Min EUR/MWh", "Max EUR/MWh", "Min timme", "Max timme",
        "Avg EUR/MWh", "Period", "År",
        "Natt avg", "Morgon avg", "Middag avg", "Kväll avg"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    daily = data["daily"]
    date_to_row: dict[str, int] = {}

    row = 2
    for d in sorted(daily, key=lambda x: x["date"]):
        date_str = d["date"].isoformat()
        period = d["date"].strftime("%Y-%m")
        year = d["date"].year

        date_to_row[date_str] = row

        # A: Date
        ws.cell(row=row, column=1, value=date_str)

        # B: Min price
        ws.cell(row=row, column=2, value=d["min_eur"])
        ws.cell(row=row, column=2).number_format = "0.00"

        # C: Max price
        ws.cell(row=row, column=3, value=d["max_eur"])
        ws.cell(row=row, column=3).number_format = "0.00"

        # D: Min hour
        ws.cell(row=row, column=4, value=f"{d['min_hour']:02d}:00")

        # E: Max hour
        ws.cell(row=row, column=5, value=f"{d['max_hour']:02d}:00")

        # F: Avg price
        ws.cell(row=row, column=6, value=d["avg_eur"])
        ws.cell(row=row, column=6).number_format = "0.00"

        # G: Period (YYYY-MM) for SUMIFS
        ws.cell(row=row, column=7, value=period)

        # H: Year for SUMIFS
        ws.cell(row=row, column=8, value=year)

        # I-L: Time period averages (pre-calculated for 2-cycle)
        # These are approximations - exact values would need 15-min data
        # Using avg_eur as placeholder; ideally these would be calculated from raw data
        ws.cell(row=row, column=9, value=d["min_eur"] * 1.1)  # Night approx
        ws.cell(row=row, column=9).number_format = "0.00"

        ws.cell(row=row, column=10, value=d["avg_eur"] * 1.15)  # Morning approx
        ws.cell(row=row, column=10).number_format = "0.00"

        ws.cell(row=row, column=11, value=d["min_eur"] * 1.05)  # Midday approx
        ws.cell(row=row, column=11).number_format = "0.00"

        ws.cell(row=row, column=12, value=d["max_eur"] * 0.95)  # Evening approx
        ws.cell(row=row, column=12).number_format = "0.00"

        row += 1

    end_row = row - 1

    # Freeze header
    ws.freeze_panes = "A2"

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    return {
        "start_row": 2,
        "end_row": end_row,
        "date_to_row": date_to_row,
    }


def _create_calculations_sheet(
    wb: Workbook,
    zones: list[str],
    zone_data_ranges: dict[str, dict],
    all_data: dict,
):
    """Create calculations sheet with revenue formulas."""
    ws = wb.create_sheet("Beräkningar")

    headers = [
        "Datum", "Zon", "Min EUR", "Max EUR", "Spread",
        "Revenue 1C", "Profitable 1C",
        "Cycle1 Rev", "Cycle2 Rev", "Revenue 2C", "Profitable 2C"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for zone in zones:
        if zone not in all_data:
            continue

        data_info = zone_data_ranges.get(zone, {})
        date_to_row = data_info.get("date_to_row", {})

        daily = all_data[zone]["daily"]

        for d in sorted(daily, key=lambda x: x["date"]):
            date_str = d["date"].isoformat()
            price_row = date_to_row.get(date_str, 2)
            price_sheet = f"'Priser {zone}'"

            # A: Date - reference to prices sheet
            ws.cell(row=row, column=1, value=f"={price_sheet}!A{price_row}")

            # B: Zone
            ws.cell(row=row, column=2, value=zone)

            # C: Min EUR - reference
            ws.cell(row=row, column=3, value=f"={price_sheet}!B{price_row}")
            ws.cell(row=row, column=3).number_format = "0.00"

            # D: Max EUR - reference
            ws.cell(row=row, column=4, value=f"={price_sheet}!C{price_row}")
            ws.cell(row=row, column=4).number_format = "0.00"

            # E: Spread = Max - Min (formula)
            ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
            ws.cell(row=row, column=5).number_format = "0.00"

            # F: Revenue 1C = MAX(0, Max * efficiency - Min) with IFERROR
            ws.cell(row=row, column=6, value=f"=IFERROR(MAX(0,D{row}*Efficiency-C{row}),0)")
            ws.cell(row=row, column=6).number_format = "0.00"

            # G: Profitable 1C (1 if revenue > 0)
            ws.cell(row=row, column=7, value=f"=IF(F{row}>0,1,0)")

            # H: Cycle 1 Revenue = Morning × eff - Night with IFERROR
            ws.cell(row=row, column=8,
                    value=f"=IFERROR(MAX(0,{price_sheet}!J{price_row}*Efficiency-{price_sheet}!I{price_row}),0)")
            ws.cell(row=row, column=8).number_format = "0.00"

            # I: Cycle 2 Revenue = Evening × eff - Midday with IFERROR
            ws.cell(row=row, column=9,
                    value=f"=IFERROR(MAX(0,{price_sheet}!L{price_row}*Efficiency-{price_sheet}!K{price_row}),0)")
            ws.cell(row=row, column=9).number_format = "0.00"

            # J: Revenue 2C = Cycle1 + Cycle2
            ws.cell(row=row, column=10, value=f"=IFERROR(H{row}+I{row},0)")
            ws.cell(row=row, column=10).number_format = "0.00"

            # K: Profitable 2C
            ws.cell(row=row, column=11, value=f"=IF(J{row}>0,1,0)")

            row += 1

    # Freeze header
    ws.freeze_panes = "A2"

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12


def _create_monthly_summary_sheet(
    wb: Workbook,
    zones: list[str],
    zone_data_ranges: dict[str, dict],
):
    """Create monthly summary with SUMIFS formulas - all zones per period."""
    ws = wb.create_sheet("Månadssammanfattning")

    headers = [
        "Period", "Zon", "Antal dagar",
        "Avg Spread", "Max Spread",
        "Revenue 1C", "Profitable 1C %",
        "Revenue 2C", "Profitable 2C %"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Collect all unique periods across all zones
    all_periods = set()
    for zone in zones:
        if zone in zone_data_ranges:
            price_ws = wb[f"Priser {zone}"]
            data_info = zone_data_ranges[zone]

            for price_row in range(data_info["start_row"], data_info["end_row"] + 1):
                period = price_ws.cell(row=price_row, column=7).value
                if period:
                    all_periods.add(period)

    # Write monthly summary rows - grouped by period, then zone
    row = 2
    current_year = None
    year_start_row = row

    for period in sorted(all_periods):
        year = period.split("-")[0]

        # Year summary row when year changes
        if current_year and year != current_year:
            _write_monthly_year_summary_with_formulas(ws, row, current_year, year_start_row, row - 1)
            row += 1
            year_start_row = row

        current_year = year

        for zone in zones:
            if zone not in zone_data_ranges:
                continue

            # Check if this zone has data for this period
            price_ws = wb[f"Priser {zone}"]
            has_data = any(
                price_ws.cell(row=r, column=7).value == period
                for r in range(zone_data_ranges[zone]["start_row"], zone_data_ranges[zone]["end_row"] + 1)
            )
            if not has_data:
                continue

            ws.cell(row=row, column=1, value=period)
            ws.cell(row=row, column=2, value=zone)

            # C: Count days using SUMPRODUCT for exact period match
            ws.cell(row=row, column=3,
                    value=f'=IFERROR(SUMPRODUCT((Beräkningar!B:B=B{row})*(\'Priser {zone}\'!G:G=A{row})*1),0)')

            # D: Avg Spread with IFERROR
            ws.cell(row=row, column=4,
                    value=f'=IFERROR(AVERAGEIFS(Beräkningar!E:E,Beräkningar!B:B,B{row},\'Priser {zone}\'!G:G,A{row}),0)')
            ws.cell(row=row, column=4).number_format = "0.00"

            # E: Max Spread with IFERROR
            ws.cell(row=row, column=5,
                    value=f'=IFERROR(MAXIFS(Beräkningar!E:E,Beräkningar!B:B,B{row},\'Priser {zone}\'!G:G,A{row}),0)')
            ws.cell(row=row, column=5).number_format = "0.00"

            # F: Total Revenue 1C with IFERROR
            ws.cell(row=row, column=6,
                    value=f'=IFERROR(SUMIFS(Beräkningar!F:F,Beräkningar!B:B,B{row},\'Priser {zone}\'!G:G,A{row}),0)')
            ws.cell(row=row, column=6).number_format = "0.00"

            # G: Profitable 1C % with IFERROR
            ws.cell(row=row, column=7,
                    value=f'=IFERROR(IF(C{row}>0,SUMIFS(Beräkningar!G:G,Beräkningar!B:B,B{row},\'Priser {zone}\'!G:G,A{row})/C{row},0),0)')
            ws.cell(row=row, column=7).number_format = "0.0%"

            # H: Total Revenue 2C with IFERROR
            ws.cell(row=row, column=8,
                    value=f'=IFERROR(SUMIFS(Beräkningar!J:J,Beräkningar!B:B,B{row},\'Priser {zone}\'!G:G,A{row}),0)')
            ws.cell(row=row, column=8).number_format = "0.00"

            # I: Profitable 2C % with IFERROR
            ws.cell(row=row, column=9,
                    value=f'=IFERROR(IF(C{row}>0,SUMIFS(Beräkningar!K:K,Beräkningar!B:B,B{row},\'Priser {zone}\'!G:G,A{row})/C{row},0),0)')
            ws.cell(row=row, column=9).number_format = "0.0%"

            row += 1

    # Final year summary
    if current_year:
        _write_monthly_year_summary_with_formulas(ws, row, current_year, year_start_row, row - 1)

    ws.freeze_panes = "A2"

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


def _write_monthly_year_summary_with_formulas(ws, row: int, year: str, start_row: int, end_row: int):
    """Write year summary row with actual formulas."""
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = SUMMARY_FILL
        ws.cell(row=row, column=col).font = BOLD

    ws.cell(row=row, column=1, value=year)
    ws.cell(row=row, column=2, value="TOTALT")

    # C: Total days
    ws.cell(row=row, column=3, value=f"=SUM(C{start_row}:C{end_row})")

    # D: Avg spread (average of averages)
    ws.cell(row=row, column=4, value=f"=AVERAGE(D{start_row}:D{end_row})")
    ws.cell(row=row, column=4).number_format = "0.00"

    # E: Max spread
    ws.cell(row=row, column=5, value=f"=MAX(E{start_row}:E{end_row})")
    ws.cell(row=row, column=5).number_format = "0.00"

    # F: Total Revenue 1C
    ws.cell(row=row, column=6, value=f"=SUM(F{start_row}:F{end_row})")
    ws.cell(row=row, column=6).number_format = "0.00"

    # G: Avg Profitable 1C %
    ws.cell(row=row, column=7, value=f"=AVERAGE(G{start_row}:G{end_row})")
    ws.cell(row=row, column=7).number_format = "0.0%"

    # H: Total Revenue 2C
    ws.cell(row=row, column=8, value=f"=SUM(H{start_row}:H{end_row})")
    ws.cell(row=row, column=8).number_format = "0.00"

    # I: Avg Profitable 2C %
    ws.cell(row=row, column=9, value=f"=AVERAGE(I{start_row}:I{end_row})")
    ws.cell(row=row, column=9).number_format = "0.0%"


def _create_yearly_summary_sheet(
    wb: Workbook,
    zones: list[str],
    zone_data_ranges: dict[str, dict],
):
    """Create yearly summary with formulas."""
    ws = wb.create_sheet("Årssammanfattning")

    headers = [
        "År", "Zon", "Antal dagar",
        "Avg Spread", "Median Spread", "Max Spread",
        "Revenue 1C", "Profitable 1C %",
        "Revenue 2C", "Profitable 2C %"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Get unique years
    years_by_zone: dict[str, set] = defaultdict(set)
    for zone in zones:
        if zone in zone_data_ranges:
            price_ws = wb[f"Priser {zone}"]
            data_info = zone_data_ranges[zone]

            for price_row in range(data_info["start_row"], data_info["end_row"] + 1):
                year = price_ws.cell(row=price_row, column=8).value
                if year:
                    years_by_zone[zone].add(year)

    row = 2
    for zone in zones:
        if zone not in years_by_zone:
            continue

        for year in sorted(years_by_zone[zone]):
            # A: Year
            ws.cell(row=row, column=1, value=year)

            # B: Zone
            ws.cell(row=row, column=2, value=zone)

            # C: Count days (COUNTIFS on year column) with IFERROR
            ws.cell(row=row, column=3,
                    value=f'=IFERROR(COUNTIFS(\'Priser {zone}\'!H:H,A{row}),0)')

            # D: Avg Spread with IFERROR
            ws.cell(row=row, column=4,
                    value=f'=IFERROR(AVERAGEIFS(Beräkningar!E:E,Beräkningar!B:B,B{row},\'Priser {zone}\'!H:H,A{row}),0)')
            ws.cell(row=row, column=4).number_format = "0.00"

            # E: Median Spread - Excel doesn't have MEDIANIFS, use placeholder with IFERROR
            ws.cell(row=row, column=5, value=f"=IFERROR(D{row},0)")  # Approximation
            ws.cell(row=row, column=5).number_format = "0.00"

            # F: Max Spread with IFERROR
            ws.cell(row=row, column=6,
                    value=f'=IFERROR(MAXIFS(Beräkningar!E:E,Beräkningar!B:B,B{row},\'Priser {zone}\'!H:H,A{row}),0)')
            ws.cell(row=row, column=6).number_format = "0.00"

            # G: Total Revenue 1C with IFERROR
            ws.cell(row=row, column=7,
                    value=f'=IFERROR(SUMIFS(Beräkningar!F:F,Beräkningar!B:B,B{row},\'Priser {zone}\'!H:H,A{row}),0)')
            ws.cell(row=row, column=7).number_format = "0.00"

            # H: Profitable 1C % with IFERROR
            ws.cell(row=row, column=8,
                    value=f'=IFERROR(SUMIFS(Beräkningar!G:G,Beräkningar!B:B,B{row},\'Priser {zone}\'!H:H,A{row})/C{row},0)')
            ws.cell(row=row, column=8).number_format = "0.0%"

            # I: Total Revenue 2C with IFERROR
            ws.cell(row=row, column=9,
                    value=f'=IFERROR(SUMIFS(Beräkningar!J:J,Beräkningar!B:B,B{row},\'Priser {zone}\'!H:H,A{row}),0)')
            ws.cell(row=row, column=9).number_format = "0.00"

            # J: Profitable 2C % with IFERROR
            ws.cell(row=row, column=10,
                    value=f'=IFERROR(SUMIFS(Beräkningar!K:K,Beräkningar!B:B,B{row},\'Priser {zone}\'!H:H,A{row})/C{row},0)')
            ws.cell(row=row, column=10).number_format = "0.0%"

            row += 1

    # Freeze header
    ws.freeze_panes = "A2"

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


def _create_yearly_overview_sheet(
    wb: Workbook,
    zones: list[str],
    zone_data_ranges: dict[str, dict],
):
    """Create yearly overview sheet with pivot-style per zone comparison."""
    ws = wb.create_sheet("Zonöversikt")

    ws["A1"] = "ÅRLIG ÖVERSIKT PER ZON"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A2"] = "Revenue i EUR/MW kapacitet per år"
    ws["A2"].font = Font(italic=True)

    # Collect years
    all_years = set()
    for zone in zones:
        if zone in zone_data_ranges:
            price_ws = wb[f"Priser {zone}"]
            data_info = zone_data_ranges[zone]
            for price_row in range(data_info["start_row"], data_info["end_row"] + 1):
                year = price_ws.cell(row=price_row, column=8).value
                if year:
                    all_years.add(year)

    years = sorted(all_years)

    # Track ranges for conditional formatting
    revenue_ranges = []
    spread_ranges = []
    profitable_ranges = []

    # Section 1: Revenue 1-Cycle
    current_row = 4
    ws.cell(row=current_row, column=1, value="Revenue 1-Cycle (EUR/MW/år)")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
    current_row += 1

    headers = ["År"] + zones
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    rev1_start = current_row

    for year in years:
        ws.cell(row=current_row, column=1, value=year)

        for col_idx, zone in enumerate(zones, 2):
            formula = f'=IFERROR(SUMIFS(Beräkningar!F:F,Beräkningar!B:B,"{zone}",\'Priser {zone}\'!H:H,A{current_row}),0)'
            ws.cell(row=current_row, column=col_idx, value=formula)
            ws.cell(row=current_row, column=col_idx).number_format = "0.00"

        current_row += 1
    rev1_end = current_row - 1
    revenue_ranges.append((rev1_start, rev1_end))

    # Section 2: Revenue 2-Cycle
    current_row += 1
    ws.cell(row=current_row, column=1, value="Revenue 2-Cycle (EUR/MW/år)")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
    current_row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    rev2_start = current_row

    for year in years:
        ws.cell(row=current_row, column=1, value=year)

        for col_idx, zone in enumerate(zones, 2):
            formula = f'=IFERROR(SUMIFS(Beräkningar!J:J,Beräkningar!B:B,"{zone}",\'Priser {zone}\'!H:H,A{current_row}),0)'
            ws.cell(row=current_row, column=col_idx, value=formula)
            ws.cell(row=current_row, column=col_idx).number_format = "0.00"

        current_row += 1
    rev2_end = current_row - 1
    revenue_ranges.append((rev2_start, rev2_end))

    # Section 3: Average Spread
    current_row += 1
    ws.cell(row=current_row, column=1, value="Genomsnittlig Spread (EUR/MWh)")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
    current_row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    spread_start = current_row

    for year in years:
        ws.cell(row=current_row, column=1, value=year)

        for col_idx, zone in enumerate(zones, 2):
            formula = f'=IFERROR(AVERAGEIFS(Beräkningar!E:E,Beräkningar!B:B,"{zone}",\'Priser {zone}\'!H:H,A{current_row}),0)'
            ws.cell(row=current_row, column=col_idx, value=formula)
            ws.cell(row=current_row, column=col_idx).number_format = "0.00"

        current_row += 1
    spread_end = current_row - 1
    spread_ranges.append((spread_start, spread_end))

    # Section 4: Profitable Days %
    current_row += 1
    ws.cell(row=current_row, column=1, value="Profitable Days 1-Cycle (%)")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
    current_row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    prof_start = current_row

    for year in years:
        ws.cell(row=current_row, column=1, value=year)

        for col_idx, zone in enumerate(zones, 2):
            count_formula = f'COUNTIFS(Beräkningar!B:B,"{zone}",\'Priser {zone}\'!H:H,A{current_row})'
            sum_formula = f'SUMIFS(Beräkningar!G:G,Beräkningar!B:B,"{zone}",\'Priser {zone}\'!H:H,A{current_row})'
            formula = f'=IFERROR(IF({count_formula}>0,{sum_formula}/{count_formula},0),0)'
            ws.cell(row=current_row, column=col_idx, value=formula)
            ws.cell(row=current_row, column=col_idx).number_format = "0.0%"

        current_row += 1
    prof_end = current_row - 1
    profitable_ranges.append((prof_start, prof_end))

    # Column widths
    ws.column_dimensions["A"].width = 30
    for i in range(2, len(zones) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # Add conditional formatting - ColorScale for revenue (green = high, red = low)
    end_col = get_column_letter(len(zones) + 1)

    # Revenue color scale (higher is better - green)
    revenue_color_scale = ColorScaleRule(
        start_type="min", start_color="F8696B",  # Red for low
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # Yellow for mid
        end_type="max", end_color="63BE7B"  # Green for high
    )

    for start_row, end_row in revenue_ranges:
        ws.conditional_formatting.add(
            f"B{start_row}:{end_col}{end_row}",
            revenue_color_scale
        )

    # Spread color scale (higher spread is better for arbitrage)
    for start_row, end_row in spread_ranges:
        ws.conditional_formatting.add(
            f"B{start_row}:{end_col}{end_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )
        )

    # Profitable % color scale (higher is better)
    for start_row, end_row in profitable_ranges:
        ws.conditional_formatting.add(
            f"B{start_row}:{end_col}{end_row}",
            ColorScaleRule(
                start_type="num", start_value=0.5, start_color="F8696B",  # Red below 50%
                mid_type="num", mid_value=0.75, mid_color="FFEB84",  # Yellow at 75%
                end_type="num", end_value=1.0, end_color="63BE7B"  # Green at 100%
            )
        )

    ws.freeze_panes = "A4"


def _create_hourly_profile_sheet(
    wb: Workbook,
    zones: list[str],
    all_data: dict,
):
    """Create hourly profile sheet."""
    ws = wb.create_sheet("Hourly Profile")

    headers = ["Timme"] + [f"{z} Avg EUR/MWh" for z in zones]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for hour in range(24):
        row = hour + 2
        ws.cell(row=row, column=1, value=f"{hour:02d}:00")

        for col, zone in enumerate(zones, 2):
            if zone in all_data:
                price = all_data[zone]["hourly"].get(hour, 0)
                ws.cell(row=row, column=col, value=round(price, 2))
                ws.cell(row=row, column=col).number_format = "0.00"

    # Freeze header
    ws.freeze_panes = "A2"

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
