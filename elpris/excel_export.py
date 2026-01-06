"""Excel export for capture price data with full formula transparency.

This module exports capture price calculations to Excel with all formulas visible,
allowing users to verify and modify calculations directly in Excel.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

from .capture import read_price_data
from .capture_report import REPORTS_DIR
from .config import ZONES
from .solar_profile import get_quarterly_solar_weight, list_available_profiles

# Default exchange rate SEK/EUR
DEFAULT_EXCHANGE_RATE = 11.50

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
BOLD = Font(bold=True)


def export_capture_excel(
    filename: str | None = None,
    exchange_rate: float = DEFAULT_EXCHANGE_RATE,
    profiles: list[str] | None = None,
    zones: list[str] | None = None,
) -> Path:
    """
    Export capture price comparison to Excel with full formula transparency.

    Creates an Excel workbook with:
    - Parametrar: Editable exchange rate
    - Data sheets per zone: Raw price data with solar profile weights
    - Beräkningar: Formulas calculating capture prices per period
    - Sammanfattning: Summary view with references to calculations

    Args:
        filename: Output filename (default: timestamped)
        exchange_rate: Default SEK/EUR exchange rate
        profiles: List of solar profiles (default: all except 'sweden')
        zones: List of zones (default: SE1-SE4)

    Returns:
        Path to exported file
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"capture_prices_{timestamp}.xlsx"

    if profiles is None:
        profiles = [p for p in list_available_profiles() if p != "sweden"]

    if zones is None:
        zones = list(ZONES)

    output_path = REPORTS_DIR / filename

    wb = Workbook()
    wb.remove(wb.active)

    # Track data row ranges for each zone (for formulas)
    zone_data_ranges: dict[str, dict] = {}

    # Sheet 1: Parameters
    _create_parameters_sheet(wb, exchange_rate, profiles, zones)

    # Sheets 2-5: Data per zone
    for zone in zones:
        data_info = _create_zone_data_sheet(wb, zone, profiles)
        zone_data_ranges[zone] = data_info

    # Sheet 6: Calculations with formulas
    _create_calculations_sheet(wb, zones, profiles, zone_data_ranges)

    # Sheet 7: Monthly summary per zone
    _create_summary_sheet(wb, zones, profiles)

    # Sheet 8: Yearly overview per zone (pivot-style)
    _create_yearly_overview_sheet(wb, zones, profiles, zone_data_ranges)

    wb.save(output_path)
    return output_path


def _create_parameters_sheet(
    wb: Workbook,
    exchange_rate: float,
    profiles: list[str],
    zones: list[str],
):
    """Create parameters sheet with editable exchange rate and best practices."""
    ws = wb.create_sheet("Parametrar")

    ws["A1"] = "CAPTURE PRICE BERÄKNING"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    # Editable parameters section
    ws["A3"] = "REDIGERBARA PARAMETRAR"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL

    ws["A4"] = "Växelkurs SEK/EUR"
    ws["B4"] = exchange_rate
    ws["B4"].fill = INPUT_FILL
    ws["B4"].number_format = "0.00"
    ws["B4"].protection = Protection(locked=False)  # Allow editing
    ws["C4"] = "(Ändra denna för att uppdatera alla EUR-beräkningar)"

    # Add comment for documentation
    ws["B4"].comment = Comment(
        "Växelkurs för konvertering SEK → EUR.\n"
        "Ändra värdet för att uppdatera alla EUR-beräkningar automatiskt.\n"
        "Typiskt värde: 11.00-12.00",
        "System"
    )

    # Data validation: exchange rate should be between 5 and 20
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1="5",
        formula2="20",
        showErrorMessage=True,
        errorTitle="Ogiltigt värde",
        error="Växelkursen måste vara mellan 5 och 20 SEK/EUR"
    )
    dv.add(ws["B4"])
    ws.add_data_validation(dv)

    # Create named range for exchange rate
    wb.defined_names.add(DefinedName("ExchangeRate", attr_text="Parametrar!$B$4"))

    # Fixed parameters
    ws["A6"] = "FASTA PARAMETRAR"
    ws["A6"].font = BOLD

    ws["A7"] = "Zoner"
    ws["B7"] = ", ".join(zones)

    ws["A8"] = "Solprofiler"
    ws["B8"] = ", ".join(profiles)

    # Legend
    ws["A10"] = "FORMELFÖRKLARING"
    ws["A10"].font = BOLD

    ws["A11"] = "Capture Price"
    ws["B11"] = "= SUMPRODUCT(pris × solvikt) / SUM(solvikt)"

    ws["A12"] = "Capture Ratio"
    ws["B12"] = "= Capture Price / Baseload Price"

    ws["A13"] = "Konvertering"
    ws["B13"] = "SEK/kWh × 1000 / växelkurs = EUR/MWh"

    # Add comment explaining the sheets
    ws["A15"] = "SHEET-STRUKTUR"
    ws["A15"].font = BOLD

    ws["A16"] = "Data SEn"
    ws["B16"] = "Rådata med priser och solprofilvikter (formel: pris × vikt)"
    ws["A17"] = "Beräkningar"
    ws["B17"] = "Aggregerade capture price-beräkningar per månad/zon"
    ws["A18"] = "Sammanfattning"
    ws["B18"] = "Månadsvy med årssnitt"
    ws["A19"] = "Årsöversikt"
    ws["B19"] = "Pivot-vy: År × Zon för snabb jämförelse"

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 50

    # Protect sheet but allow editing of input cells (no password)
    ws.protection.sheet = True
    ws.protection.enable()


def _create_zone_data_sheet(
    wb: Workbook,
    zone: str,
    profiles: list[str],
) -> dict:
    """
    Create data sheet for a zone with raw price data and solar weights.

    Returns dict with start_row, end_row, and period_rows for formulas.
    """
    ws = wb.create_sheet(f"Data {zone}")

    # Headers
    headers = ["Timestamp", "Period", "Pris SEK/kWh"]
    for profile in profiles:
        headers.append(f"{_display_name(profile)} vikt")
        headers.append(f"{_display_name(profile)} pris×vikt")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Read all price data for this zone
    price_data = list(read_price_data(zone))

    if not price_data:
        ws["A2"] = "Ingen data tillgänglig"
        return {"start_row": 2, "end_row": 2, "period_rows": {}}

    # Track periods for aggregation formulas
    period_rows: dict[str, list[int]] = {}

    row = 2
    for record in price_data:
        ts = record["_timestamp"]
        period = ts.strftime("%Y-%m")
        price_sek = float(record["SEK_per_kWh"])

        # Track rows per period
        if period not in period_rows:
            period_rows[period] = []
        period_rows[period].append(row)

        # Column A: Timestamp
        ws.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d %H:%M"))

        # Column B: Period
        ws.cell(row=row, column=2, value=period)

        # Column C: Price SEK/kWh
        ws.cell(row=row, column=3, value=price_sek)
        ws.cell(row=row, column=3).number_format = "0.0000"

        # Columns D onwards: Solar weights and weighted prices
        col = 4
        for profile in profiles:
            weight = get_quarterly_solar_weight(ts, profile)

            # Weight column
            ws.cell(row=row, column=col, value=weight)
            ws.cell(row=row, column=col).number_format = "0.0000"

            # Weighted price formula: =C{row}*D{row}
            weight_col = get_column_letter(col)
            ws.cell(row=row, column=col + 1, value=f"=C{row}*{weight_col}{row}")
            ws.cell(row=row, column=col + 1).number_format = "0.000000"

            col += 2

        row += 1

    end_row = row - 1

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    for i in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    return {
        "start_row": 2,
        "end_row": end_row,
        "period_rows": period_rows,
    }


def _create_calculations_sheet(
    wb: Workbook,
    zones: list[str],
    profiles: list[str],
    zone_data_ranges: dict[str, dict],
):
    """Create calculations sheet with SUMPRODUCT formulas."""
    ws = wb.create_sheet("Beräkningar")

    # Headers
    headers = ["Period", "Zon"]
    for profile in profiles:
        name = _display_name(profile)
        headers.extend([
            f"{name} Σ(pris×vikt)",
            f"{name} Σ(vikt)",
            f"{name} Capture SEK",
            f"{name} Capture EUR",
            f"{name} Ratio",
        ])
    headers.extend(["Baseload SEK", "Baseload EUR"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Generate one row per zone per period
    row = 2

    for zone in zones:
        data_info = zone_data_ranges.get(zone, {})
        period_rows = data_info.get("period_rows", {})

        for period in sorted(period_rows.keys()):
            rows = period_rows[period]
            first_row = min(rows)
            last_row = max(rows)

            ws.cell(row=row, column=1, value=period)
            ws.cell(row=row, column=2, value=zone)

            col = 3
            for i, profile in enumerate(profiles):
                # Column indices in Data sheet:
                # C = price, D = weight1, E = weighted1, F = weight2, G = weighted2, etc.
                weight_col = get_column_letter(4 + i * 2)  # D, F, H, ...
                weighted_col = get_column_letter(5 + i * 2)  # E, G, I, ...

                data_sheet = f"'Data {zone}'"

                # Sum of (price × weight)
                sum_weighted_formula = f"=SUM({data_sheet}!{weighted_col}{first_row}:{weighted_col}{last_row})"
                ws.cell(row=row, column=col, value=sum_weighted_formula)
                ws.cell(row=row, column=col).number_format = "0.0000"

                # Sum of weights
                sum_weight_formula = f"=SUM({data_sheet}!{weight_col}{first_row}:{weight_col}{last_row})"
                ws.cell(row=row, column=col + 1, value=sum_weight_formula)
                ws.cell(row=row, column=col + 1).number_format = "0.0000"

                # Capture SEK = sum_weighted / sum_weight (with IFERROR)
                sum_weighted_col = get_column_letter(col)
                sum_weight_col = get_column_letter(col + 1)
                capture_sek_formula = f"=IFERROR(IF({sum_weight_col}{row}>0,{sum_weighted_col}{row}/{sum_weight_col}{row},0),0)"
                ws.cell(row=row, column=col + 2, value=capture_sek_formula)
                ws.cell(row=row, column=col + 2).number_format = "0.0000"

                # Capture EUR = Capture_SEK * 1000 / ExchangeRate (using named range)
                capture_sek_col = get_column_letter(col + 2)
                capture_eur_formula = f"=IFERROR({capture_sek_col}{row}*1000/ExchangeRate,0)"
                ws.cell(row=row, column=col + 3, value=capture_eur_formula)
                ws.cell(row=row, column=col + 3).number_format = "0.00"

                col += 5

            # Baseload SEK = AVERAGE(price) with IFERROR
            baseload_sek_formula = f"=IFERROR(AVERAGE({data_sheet}!C{first_row}:C{last_row}),0)"
            ws.cell(row=row, column=col, value=baseload_sek_formula)
            ws.cell(row=row, column=col).number_format = "0.0000"

            # Baseload EUR using named range
            baseload_sek_col = get_column_letter(col)
            baseload_eur_formula = f"=IFERROR({baseload_sek_col}{row}*1000/ExchangeRate,0)"
            ws.cell(row=row, column=col + 1, value=baseload_eur_formula)
            ws.cell(row=row, column=col + 1).number_format = "0.00"

            # Calculate ratio for each profile (with IFERROR)
            baseload_eur_col = get_column_letter(col + 1)
            for i, profile in enumerate(profiles):
                capture_eur_col = get_column_letter(3 + i * 5 + 3)  # Column with Capture EUR
                ratio_col_idx = 3 + i * 5 + 4  # Column with Ratio

                ratio_formula = f"=IFERROR(IF({baseload_eur_col}{row}>0,{capture_eur_col}{row}/{baseload_eur_col}{row},0),0)"
                ws.cell(row=row, column=ratio_col_idx, value=ratio_formula)
                ws.cell(row=row, column=ratio_col_idx).number_format = "0.0%"

            row += 1

    # Freeze header
    ws.freeze_panes = "A2"

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


def _create_summary_sheet(
    wb: Workbook,
    zones: list[str],
    profiles: list[str],
):
    """Create summary sheet with references to calculations."""
    ws = wb.create_sheet("Sammanfattning")

    # Title
    ws["A1"] = "CAPTURE PRICE SAMMANFATTNING"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A2"] = "Alla värden i EUR/MWh. Capture Ratio visar capture price som andel av baseload."
    ws["A2"].font = Font(italic=True)

    # Headers
    headers = ["Period", "Zon"]
    for profile in profiles:
        name = _display_name(profile)
        headers.extend([f"{name} EUR/MWh", f"{name} %"])
    headers.append("Baseload EUR/MWh")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    calc_ws = wb["Beräkningar"]
    calc_row_count = calc_ws.max_row

    # Build row mapping: calc_row -> summary_row for AVERAGE formulas
    row = 5
    year_start_rows: dict[str, int] = {}  # year -> first summary row
    year_end_rows: dict[str, int] = {}    # year -> last summary row
    current_year = None

    for calc_row in range(2, calc_row_count + 1):
        period_cell = calc_ws.cell(row=calc_row, column=1).value
        if not period_cell:
            continue

        year = period_cell.split("-")[0]

        # Track year boundaries for AVERAGE formulas
        if year not in year_start_rows:
            year_start_rows[year] = row

        # Insert year summary when year changes
        if current_year and year != current_year:
            year_end_rows[current_year] = row - 1
            _write_summary_year_row_with_formulas(
                ws, row, current_year, profiles,
                year_start_rows[current_year], row - 1
            )
            row += 1
            year_start_rows[year] = row

        current_year = year

        # Write data row with references
        ws.cell(row=row, column=1, value=f"=Beräkningar!A{calc_row}")
        ws.cell(row=row, column=2, value=f"=Beräkningar!B{calc_row}")

        col = 3
        for i, profile in enumerate(profiles):
            capture_eur_col = get_column_letter(6 + i * 5)
            ratio_col = get_column_letter(7 + i * 5)

            ws.cell(row=row, column=col, value=f"=Beräkningar!{capture_eur_col}{calc_row}")
            ws.cell(row=row, column=col).number_format = "0.00"

            ws.cell(row=row, column=col + 1, value=f"=Beräkningar!{ratio_col}{calc_row}")
            ws.cell(row=row, column=col + 1).number_format = "0.0%"

            col += 2

        num_profile_cols = len(profiles) * 5
        baseload_eur_col = get_column_letter(3 + num_profile_cols + 1)
        ws.cell(row=row, column=col, value=f"=Beräkningar!{baseload_eur_col}{calc_row}")
        ws.cell(row=row, column=col).number_format = "0.00"

        row += 1

    # Final year summary
    if current_year:
        year_end_rows[current_year] = row - 1
        _write_summary_year_row_with_formulas(
            ws, row, current_year, profiles,
            year_start_rows[current_year], row - 1
        )

    ws.freeze_panes = "A5"

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def _write_summary_year_row_with_formulas(
    ws, row: int, year: str, profiles: list[str],
    start_row: int, end_row: int
):
    """Write a year summary row with AVERAGE formulas."""
    ws.cell(row=row, column=1, value=year)
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=1).fill = SUMMARY_FILL

    ws.cell(row=row, column=2, value="SNITT")
    ws.cell(row=row, column=2).font = BOLD
    ws.cell(row=row, column=2).fill = SUMMARY_FILL

    col = 3
    for profile in profiles:
        # Average of capture EUR
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})")
        ws.cell(row=row, column=col).number_format = "0.00"
        ws.cell(row=row, column=col).font = BOLD
        ws.cell(row=row, column=col).fill = SUMMARY_FILL

        # Average of ratio
        col_letter = get_column_letter(col + 1)
        ws.cell(row=row, column=col + 1, value=f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})")
        ws.cell(row=row, column=col + 1).number_format = "0.0%"
        ws.cell(row=row, column=col + 1).font = BOLD
        ws.cell(row=row, column=col + 1).fill = SUMMARY_FILL

        col += 2

    # Average baseload
    col_letter = get_column_letter(col)
    ws.cell(row=row, column=col, value=f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})")
    ws.cell(row=row, column=col).number_format = "0.00"
    ws.cell(row=row, column=col).font = BOLD
    ws.cell(row=row, column=col).fill = SUMMARY_FILL


def _create_yearly_overview_sheet(
    wb: Workbook,
    zones: list[str],
    profiles: list[str],
    zone_data_ranges: dict[str, dict],
):
    """Create yearly overview sheet with pivot-style summary per zone."""
    ws = wb.create_sheet("Årsöversikt")

    # Title
    ws["A1"] = "ÅRLIG ÖVERSIKT PER ZON"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A2"] = "Capture price och ratio per profil, aggregerat per år och zon (EUR/MWh)"
    ws["A2"].font = Font(italic=True)

    # Collect unique years from all zones
    all_years = set()
    for zone in zones:
        if zone in zone_data_ranges:
            for period in zone_data_ranges[zone].get("period_rows", {}).keys():
                all_years.add(period.split("-")[0])

    years = sorted(all_years)

    # Create one section per profile
    current_row = 4

    for profile in profiles:
        profile_name = _display_name(profile)

        # Profile header
        ws.cell(row=current_row, column=1, value=f"{profile_name} - Capture Price EUR/MWh")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
        current_row += 1

        # Headers: År, SE1, SE2, SE3, SE4
        headers = ["År"] + zones
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Data rows per year
        for year in years:
            ws.cell(row=current_row, column=1, value=int(year))

            for col_idx, zone in enumerate(zones, 2):
                # Find the calc rows for this year and zone
                # We need to use AVERAGEIFS on Beräkningar sheet
                profile_idx = profiles.index(profile)
                capture_eur_col = get_column_letter(6 + profile_idx * 5)

                # AVERAGEIFS: average capture EUR where zone matches and period starts with year
                formula = f'=AVERAGEIFS(Beräkningar!{capture_eur_col}:{capture_eur_col},Beräkningar!B:B,"{zone}",Beräkningar!A:A,"{year}-*")'
                ws.cell(row=current_row, column=col_idx, value=formula)
                ws.cell(row=current_row, column=col_idx).number_format = "0.00"

            current_row += 1

        # Add ratio section for this profile
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{profile_name} - Capture Ratio %")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
        current_row += 1

        # Headers again
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for year in years:
            ws.cell(row=current_row, column=1, value=int(year))

            for col_idx, zone in enumerate(zones, 2):
                profile_idx = profiles.index(profile)
                ratio_col = get_column_letter(7 + profile_idx * 5)

                formula = f'=AVERAGEIFS(Beräkningar!{ratio_col}:{ratio_col},Beräkningar!B:B,"{zone}",Beräkningar!A:A,"{year}-*")'
                ws.cell(row=current_row, column=col_idx, value=formula)
                ws.cell(row=current_row, column=col_idx).number_format = "0.0%"

            current_row += 1

        current_row += 2  # Space between profiles

    # Baseload overview
    ws.cell(row=current_row, column=1, value="Baseload - Genomsnittspris EUR/MWh")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(zones) + 1)
    current_row += 1

    headers = ["År"] + zones
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    num_profile_cols = len(profiles) * 5
    baseload_col = get_column_letter(3 + num_profile_cols + 1)

    for year in years:
        ws.cell(row=current_row, column=1, value=int(year))

        for col_idx, zone in enumerate(zones, 2):
            formula = f'=AVERAGEIFS(Beräkningar!{baseload_col}:{baseload_col},Beräkningar!B:B,"{zone}",Beräkningar!A:A,"{year}-*")'
            ws.cell(row=current_row, column=col_idx, value=formula)
            ws.cell(row=current_row, column=col_idx).number_format = "0.00"

        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    for i in range(2, len(zones) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = "A4"

    # Add conditional formatting for ratio columns (color scale: red < 85% < yellow < 95% < green)
    # Find ratio sections and apply color scale
    for row_idx in range(1, current_row):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and "Ratio" in str(cell_val):
            # This is a ratio header row, data starts 2 rows down
            data_start = row_idx + 2
            data_end = data_start + len(years) - 1

            for col_idx in range(2, len(zones) + 2):
                col_letter = get_column_letter(col_idx)
                range_str = f"{col_letter}{data_start}:{col_letter}{data_end}"

                # Color scale: red (low) -> yellow (mid) -> green (high)
                color_scale = ColorScaleRule(
                    start_type="num", start_value=0.7, start_color="F8696B",  # Red
                    mid_type="num", mid_value=0.9, mid_color="FFEB84",        # Yellow
                    end_type="num", end_value=1.0, end_color="63BE7B"         # Green
                )
                ws.conditional_formatting.add(range_str, color_scale)


def _display_name(profile: str) -> str:
    """Get display name for profile."""
    names = {
        "ew_boda": "E-W",
        "south_lundby": "South",
        "tracker_sweden": "Tracker",
    }
    if profile in names:
        return names[profile]
    if profile.startswith("tracker"):
        return "Tracker"
    return profile[:8]
