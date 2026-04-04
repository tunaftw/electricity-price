"""Tidsstegsbaserad simulering av batteribalansering för solprognosfel.

Simulerar hur ett batteri (t.ex. 1 MW / 1 MWh) kan absorbera obalanser
orsakade av prognosfel i solproduktion, med verkliga eSett-obalanspriser.
"""

from __future__ import annotations

import csv
import math
import random
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

from .config import QUARTERLY_DIR, RAW_DIR
from .solar_profile import load_pvsyst_profile

# Stockholm timezone offset helpers (no pytz dependency)
_CET_OFFSET = timedelta(hours=1)
_CEST_OFFSET = timedelta(hours=2)

# Sentinel for missing price data
_NO_PRICE = float("nan")


def _stockholm_offset(dt_utc: datetime) -> timedelta:
    """Return UTC offset for Europe/Stockholm at given UTC datetime."""
    year = dt_utc.year
    # Last Sunday in March
    d = datetime(year, 3, 31, tzinfo=timezone.utc)
    march_sun = d - timedelta(days=(d.weekday() + 1) % 7)
    dst_start = march_sun.replace(hour=1, minute=0, second=0)

    # Last Sunday in October
    d = datetime(year, 10, 31, tzinfo=timezone.utc)
    oct_sun = d - timedelta(days=(d.weekday() + 1) % 7)
    dst_end = oct_sun.replace(hour=1, minute=0, second=0)

    dt_aware = dt_utc if dt_utc.tzinfo else dt_utc.replace(tzinfo=timezone.utc)
    if dst_start <= dt_aware < dst_end:
        return _CEST_OFFSET
    return _CET_OFFSET


# ---------------------------------------------------------------------------
# Pre-computed year grid: arrays aligned to UTC 15-min slots
# ---------------------------------------------------------------------------

@dataclass
class YearGrid:
    """Pre-computed arrays for one year, aligned to 15-min UTC grid."""
    year: int
    n_slots: int
    ts_keys: list[str]          # UTC timestamp strings
    day_indices: list[int]       # Slot index where each new day starts
    solar_per_mwp: list[float]  # Production (MW) per MWp installed
    esett_sales: list[float]    # EUR/MWh (NaN if missing)
    esett_purchase: list[float]
    spot_eur_mwh: list[float]   # EUR/MWh (NaN if missing)


def _build_utc_grid(year: int) -> tuple[list[str], list[int], list[tuple[int, int, int]]]:
    """Build UTC 15-min grid for a year.

    Returns (ts_keys, day_start_indices, local_keys) where local_keys
    are (month, day, hour) in Stockholm time for profile lookup.
    """
    start = datetime(year, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    dt = timedelta(minutes=15)

    ts_keys = []
    day_indices = []
    local_keys = []

    current = start
    prev_day = ""
    idx = 0
    while current < end:
        ts_key = current.strftime("%Y-%m-%dT%H:%M:%S")
        ts_keys.append(ts_key)

        day = ts_key[:10]
        if day != prev_day:
            day_indices.append(idx)
            prev_day = day

        # Stockholm local time for solar profile lookup
        offset = _stockholm_offset(current)
        local = current + offset
        local_keys.append((local.month, local.day, local.hour))

        current += dt
        idx += 1

    return ts_keys, day_indices, local_keys


def build_year_grid(
    zone: str,
    year: int,
    profile_name: str,
    esett_prices: dict[str, dict],
    spot_prices: dict[str, float],
) -> YearGrid:
    """Build pre-computed grid for one zone/year."""
    ts_keys, day_indices, local_keys = _build_utc_grid(year)
    n = len(ts_keys)

    # Solar profile (production per MWp)
    profile = load_pvsyst_profile(profile_name)
    solar_per_mwp = [profile.get(lk, 0.0) for lk in local_keys]

    # Align prices to grid
    esett_sales = []
    esett_purchase = []
    spot_arr = []
    for ts in ts_keys:
        ep = esett_prices.get(ts)
        if ep is not None:
            esett_sales.append(ep["sales_price"])
            esett_purchase.append(ep["purchase_price"])
        else:
            esett_sales.append(_NO_PRICE)
            esett_purchase.append(_NO_PRICE)

        sp = spot_prices.get(ts)
        spot_arr.append(sp if sp is not None else _NO_PRICE)

    return YearGrid(
        year=year,
        n_slots=n,
        ts_keys=ts_keys,
        day_indices=day_indices,
        solar_per_mwp=solar_per_mwp,
        esett_sales=esett_sales,
        esett_purchase=esett_purchase,
        spot_eur_mwh=spot_arr,
    )


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_esett_prices(zone: str, years: list[int]) -> dict[str, dict]:
    """Load eSett imbalance prices for a zone.

    Returns dict mapping UTC timestamp string to {sales_price, purchase_price}.
    """
    prices = {}
    for year in years:
        path = RAW_DIR / "esett" / "imbalance" / zone / f"{year}.csv"
        if not path.exists():
            continue
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ts_key = row["time_start"].rstrip("Z")
                prices[ts_key] = {
                    "sales_price": float(row["imbl_sales_price_eur_mwh"]),
                    "purchase_price": float(row["imbl_purchase_price_eur_mwh"]),
                }
    return prices


def load_spot_prices(zone: str, years: list[int]) -> dict[str, float]:
    """Load quarterly spot prices, converting to UTC keys. Returns EUR/MWh."""
    prices = {}
    for year in years:
        path = QUARTERLY_DIR / zone / f"{year}.csv"
        if not path.exists():
            continue
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ts_local = datetime.fromisoformat(row["time_start"])
                ts_utc = ts_local.astimezone(timezone.utc)
                ts_key = ts_utc.strftime("%Y-%m-%dT%H:%M:%S")
                prices[ts_key] = float(row["EUR_per_kWh"]) * 1000
    return prices


# ---------------------------------------------------------------------------
# Forecast error generation (AR(1) model)
# ---------------------------------------------------------------------------

def generate_forecast_errors(
    solar_mw: list[float],
    mape: float,
    seed: int,
    rho: float = 0.7,
) -> list[float]:
    """Generate correlated forecast errors using AR(1) model.

    Args:
        solar_mw: Production in MW per timestep.
        mape: Mean Absolute Percentage Error (e.g. 0.05).
        seed: Random seed.
        rho: AR(1) autocorrelation coefficient.

    Returns:
        list of forecast errors in MW.
    """
    rng = random.Random(seed)
    innovation_std = math.sqrt(1 - rho * rho) * mape * 1.4

    errors = []
    prev_rel = 0.0

    for prod in solar_mw:
        if prod < 0.001:
            errors.append(0.0)
            prev_rel = 0.0
            continue

        innovation = rng.gauss(0, innovation_std)
        rel_error = rho * prev_rel + innovation
        errors.append(rel_error * prod)
        prev_rel = rel_error

    return errors


# ---------------------------------------------------------------------------
# Simulation core (optimized inner loop)
# ---------------------------------------------------------------------------

@dataclass
class SimulationResult:
    """Result from a single year simulation."""
    zone: str
    year: int
    park_mwp: float
    mape: float
    total_cost_no_battery: float
    total_cost_with_battery: float
    total_imbalance_energy_mwh: float
    absorbed_energy_mwh: float
    residual_energy_mwh: float
    coverage_pct: float
    max_residual_mw: float
    p99_residual_mw: float
    n_timesteps: int
    n_active_timesteps: int


def simulate_year(
    zone: str,
    grid: YearGrid,
    park_mwp: float,
    battery_kw: float,
    battery_kwh: float,
    mape: float,
    seed: int,
) -> SimulationResult:
    """Run one year simulation using pre-computed grid.

    Battery parameters in kW/kWh. All internal math in kW/kWh for speed.
    """
    dt = 0.25  # hours
    eff_sqrt = math.sqrt(0.88)  # sqrt of round-trip efficiency
    soc_min = battery_kwh * 0.05
    soc_max = battery_kwh * 0.95
    soc = battery_kwh * 0.50

    # Scale solar profile to park size
    solar_mw = [s * park_mwp for s in grid.solar_per_mwp]

    # Generate forecast errors
    errors = generate_forecast_errors(solar_mw, mape, seed)

    # Pre-fetch arrays for fast indexed access
    esett_s = grid.esett_sales
    esett_p = grid.esett_purchase
    spot = grid.spot_eur_mwh
    n = grid.n_slots

    # Day boundary set for midnight SoC reset
    day_set = set(grid.day_indices)

    total_cost_no_batt = 0.0
    total_cost_with_batt = 0.0
    total_imb_e = 0.0
    absorbed_e = 0.0
    residual_e = 0.0
    max_residual = 0.0
    residual_abs_list = []
    n_active = 0

    for i in range(n):
        # Reset SoC at day boundaries (skip first)
        if i in day_set and i > 0:
            soc = battery_kwh * 0.50

        err = errors[i]
        prod = solar_mw[i]

        if prod > 0.001:
            n_active += 1

        # -- Price imbalance without battery --
        sp = spot[i]
        es = esett_s[i]
        ep = esett_p[i]

        abs_err = abs(err)
        vol = abs_err * dt  # MWh

        # Check for valid prices (not NaN)
        has_price = sp == sp and es == es  # NaN != NaN

        if has_price and abs_err > 1e-6:
            if err > 0:
                c = (sp - es) * vol
            else:
                c = (ep - sp) * vol
            cost_no = c if c > 0 else 0.0
        else:
            cost_no = 0.0
        total_cost_no_batt += cost_no

        # -- Battery step (inlined for speed) --
        err_kw = err * 1000

        if err > 0:
            # Charge: over-production
            charge_kw = min(battery_kw, err_kw)
            avail_cap = soc_max - soc
            max_e = avail_cap / eff_sqrt if eff_sqrt > 0 else 0.0
            actual_e = min(charge_kw * dt, max_e)
            actual_kw = actual_e / dt
            soc += actual_e * eff_sqrt
            absorbed_mw = actual_kw / 1000
            residual_mw = err - absorbed_mw

        elif err < 0:
            # Discharge: under-production
            discharge_kw = min(battery_kw, -err_kw)
            avail_e = soc - soc_min
            max_out = avail_e * eff_sqrt
            actual_out = min(discharge_kw * dt, max_out)
            actual_kw = actual_out / dt
            soc -= actual_out / eff_sqrt
            absorbed_mw = -(actual_kw / 1000)
            residual_mw = err - absorbed_mw
        else:
            absorbed_mw = 0.0
            residual_mw = 0.0

        # -- Price residual with battery --
        abs_res = abs(residual_mw)
        res_vol = abs_res * dt

        if has_price and abs_res > 1e-6:
            if residual_mw > 0:
                c = (sp - es) * res_vol
            else:
                c = (ep - sp) * res_vol
            cost_with = c if c > 0 else 0.0
        else:
            cost_with = 0.0
        total_cost_with_batt += cost_with

        # -- Track volumes --
        total_imb_e += vol
        absorbed_e += abs(absorbed_mw) * dt
        residual_e += res_vol
        residual_abs_list.append(abs_res)
        if abs_res > max_residual:
            max_residual = abs_res

    coverage = (absorbed_e / total_imb_e * 100) if total_imb_e > 0 else 100.0
    p99_res = _percentile(residual_abs_list, 99)

    return SimulationResult(
        zone=zone,
        year=grid.year,
        park_mwp=park_mwp,
        mape=mape,
        total_cost_no_battery=total_cost_no_batt,
        total_cost_with_battery=total_cost_with_batt,
        total_imbalance_energy_mwh=total_imb_e,
        absorbed_energy_mwh=absorbed_e,
        residual_energy_mwh=residual_e,
        coverage_pct=coverage,
        max_residual_mw=max_residual,
        p99_residual_mw=p99_res,
        n_timesteps=n,
        n_active_timesteps=n_active,
    )


# ---------------------------------------------------------------------------
# Parametric sweep
# ---------------------------------------------------------------------------

@dataclass
class SweepResult:
    """Aggregated results for one zone/park_size/mape combination."""
    zone: str
    park_mwp: float
    mape: float
    mean_coverage_pct: float
    p01_coverage_pct: float      # 1st percentile = worst-case (P99 confidence)
    mean_cost_no_battery: float
    mean_cost_with_battery: float
    mean_savings: float
    p99_residual_mw: float


def run_analysis(
    zones: list[str],
    park_sizes: list[float],
    mapes: list[float],
    battery_mw: float,
    battery_mwh: float,
    n_sims: int,
    years: list[int],
    profile: str,
    base_seed: int = 42,
) -> dict[str, list[SweepResult]]:
    """Run full parametric sweep."""
    battery_kw = battery_mw * 1000
    battery_kwh_val = battery_mwh * 1000
    results: dict[str, list[SweepResult]] = {}

    for zone in zones:
        print(f"\nLoading data for {zone}...")
        esett_prices = load_esett_prices(zone, years)
        spot_prices = load_spot_prices(zone, years)
        print(f"  eSett: {len(esett_prices)} timesteps, Spot: {len(spot_prices)} timesteps")

        # Pre-compute grids per year (shared across all park sizes and sims)
        grids = {}
        for year in years:
            print(f"  Building grid for {year}...")
            grids[year] = build_year_grid(zone, year, profile, esett_prices, spot_prices)

        zone_results = []

        for mape in mapes:
            print(f"\n  MAPE {mape*100:.1f}%:")
            for park_mwp in park_sizes:
                coverages = []
                costs_no_batt = []
                costs_with_batt = []
                p99_residuals = []

                for year in years:
                    grid = grids[year]
                    for sim_i in range(n_sims):
                        seed = base_seed + sim_i * 1000 + year
                        result = simulate_year(
                            zone=zone,
                            grid=grid,
                            park_mwp=park_mwp,
                            battery_kw=battery_kw,
                            battery_kwh=battery_kwh_val,
                            mape=mape,
                            seed=seed,
                        )
                        coverages.append(result.coverage_pct)
                        costs_no_batt.append(result.total_cost_no_battery)
                        costs_with_batt.append(result.total_cost_with_battery)
                        p99_residuals.append(result.p99_residual_mw)

                mean_cov = sum(coverages) / len(coverages)
                p01_cov = _percentile(coverages, 1)
                mean_no = sum(costs_no_batt) / len(costs_no_batt)
                mean_with = sum(costs_with_batt) / len(costs_with_batt)
                p99_res = _percentile(p99_residuals, 99)

                sweep = SweepResult(
                    zone=zone,
                    park_mwp=park_mwp,
                    mape=mape,
                    mean_coverage_pct=mean_cov,
                    p01_coverage_pct=p01_cov,
                    mean_cost_no_battery=mean_no,
                    mean_cost_with_battery=mean_with,
                    mean_savings=mean_no - mean_with,
                    p99_residual_mw=p99_res,
                )
                zone_results.append(sweep)

                print(f"    {park_mwp:5.1f} MWp: coverage={mean_cov:.1f}% "
                      f"(P99={p01_cov:.1f}%), savings={mean_no - mean_with:.0f} EUR/yr")

        results[zone] = zone_results

    return results


def find_max_mwp(
    results: list[SweepResult],
    target_coverage: float = 99.0,
) -> dict[float, float | None]:
    """Find max MWp where P99 coverage >= target for each MAPE."""
    by_mape: dict[float, list[SweepResult]] = {}
    for r in results:
        by_mape.setdefault(r.mape, []).append(r)

    out = {}
    for mape, sweep_list in sorted(by_mape.items()):
        sweep_list.sort(key=lambda s: s.park_mwp)
        max_mwp = None
        for s in sweep_list:
            if s.p01_coverage_pct >= target_coverage:
                max_mwp = s.park_mwp
            else:
                break
        out[mape] = max_mwp
    return out


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

def format_table(
    results: dict[str, list[SweepResult]],
    battery_mw: float,
    battery_mwh: float,
    profile: str,
    n_sims: int,
    target_coverage: float = 99.0,
) -> str:
    """Format results as console table."""
    lines = []

    for zone, sweep_list in results.items():
        lines.append("")
        lines.append("=" * 75)
        lines.append(f"IMBALANCE BATTERY COVERAGE - {zone}")
        lines.append(f"Battery: {battery_mw} MW / {battery_mwh} MWh | "
                      f"Profile: {profile} | Sims: {n_sims}")
        lines.append("=" * 75)

        by_mape: dict[float, list[SweepResult]] = {}
        for r in sweep_list:
            by_mape.setdefault(r.mape, []).append(r)

        for mape in sorted(by_mape.keys()):
            sweep = sorted(by_mape[mape], key=lambda s: s.park_mwp)
            lines.append(f"\nMAPE {mape*100:.1f}%:")
            lines.append(f"{'MWp':>6} | {'Coverage P99':>12} | "
                          f"{'Cost no batt':>12} | {'Cost w/ batt':>12} | {'Savings':>10}")
            lines.append("-" * 65)

            for s in sweep:
                lines.append(
                    f"{s.park_mwp:6.1f} | {s.p01_coverage_pct:11.1f}% | "
                    f"{s.mean_cost_no_battery:10.0f} EUR | "
                    f"{s.mean_cost_with_battery:10.0f} EUR | "
                    f"{s.mean_savings:8.0f} EUR"
                )

        max_mwp = find_max_mwp(sweep_list, target_coverage)
        lines.append("")
        parts = []
        for mape, mwp in sorted(max_mwp.items()):
            if mwp is not None:
                parts.append(f"MAPE {mape*100:.0f}%: {mwp:.0f} MWp")
            else:
                parts.append(f"MAPE {mape*100:.0f}%: <1 MWp")
        lines.append(f">>> Max MWp vid P{target_coverage:.0f} coverage: " + " | ".join(parts))

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(
    results: dict[str, list[SweepResult]],
    battery_mw: float,
    battery_mwh: float,
    profile: str,
    n_sims: int,
    output_path: Path,
) -> None:
    """Export results to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    for zone_idx, (zone, sweep_list) in enumerate(results.items()):
        if zone_idx == 0:
            ws = wb.active
            ws.title = zone
        else:
            ws = wb.create_sheet(zone)

        ws.append([f"Imbalance Battery Coverage - {zone}"])
        ws.append([f"Battery: {battery_mw} MW / {battery_mwh} MWh | Profile: {profile} | Sims: {n_sims}"])
        ws.append([])

        headers = ["MAPE", "Park MWp", "Coverage P99 %", "Cost no battery (EUR/yr)",
                    "Cost with battery (EUR/yr)", "Savings (EUR/yr)", "P99 residual (MW)"]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for s in sorted(sweep_list, key=lambda x: (x.mape, x.park_mwp)):
            ws.append([
                f"{s.mape*100:.1f}%",
                s.park_mwp,
                round(s.p01_coverage_pct, 1),
                round(s.mean_cost_no_battery, 0),
                round(s.mean_cost_with_battery, 0),
                round(s.mean_savings, 0),
                round(s.p99_residual_mw, 3),
            ])

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = 20

    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _percentile(data: list[float], p: float) -> float:
    """Calculate percentile without numpy."""
    sorted_data = sorted(data)
    n = len(sorted_data)
    if n == 0:
        return 0.0
    k = (n - 1) * (p / 100)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return sorted_data[int(k)]
    return sorted_data[f] * (c - k) + sorted_data[c] * (k - f)
