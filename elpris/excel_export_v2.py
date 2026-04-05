"""Data-driven Excel export for Dashboard v2.

Generates .xlsx with one sheet per zone x granularity level.
Auto-discovers profiles from the data dict — new profiles appear as columns
without code changes.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LEVEL_ORDER = ["yearly", "monthly", "daily", "hourly"]

LEVEL_LABELS = {
    "yearly": "Årsvis",
    "monthly": "Månadsvis",
    "daily": "Daglig",
    "hourly": "Timvis",
}

LEVEL_TIME_COLS = {
    "yearly": [("year", "År")],
    "monthly": [("year", "År"), ("month", "Månad")],
    "daily": [("date", "Datum")],
    "hourly": [("date", "Datum"), ("hour", "Timme (UTC)")],
}

MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "Maj", "Jun",
    "Jul", "Aug", "Sep", "Okt", "Nov", "Dec",
]

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
NUM_FMT_PRICE = '#,##0.00'
NUM_FMT_RATIO = '0.000'
NUM_FMT_WEIGHT = '0.0000'

ZONE_TAB_COLORS = {
    "SE1": "3B82F6",
    "SE2": "10B981",
    "SE3": "F59E0B",
    "SE4": "EF4444",
}


def generate_dashboard_excel(data: dict, output_path: Path) -> None:
    """Generate Excel report from dashboard v2 data.

    Creates one sheet per zone x granularity level.
    Profiles are auto-discovered from data["profiles"].
    """
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    zones = data.get("zones", [])
    profiles = data.get("profiles", {})
    profile_meta = data.get("profile_meta", {})
    all_data = data.get("data", {})

    for zone in zones:
        zone_data = all_data.get(zone, {})
        if not zone_data:
            continue

        # Profiles available for this zone (exclude baseload and BESS)
        available_profiles = [
            k for k in profiles
            if k != "baseload"
            and k in zone_data
            and not profile_meta.get(k, {}).get("category", "").startswith("bess")
        ]

        # Which granularities exist in the data?
        sample = zone_data.get("baseload", {})
        available_levels = [lv for lv in LEVEL_ORDER if lv in sample]

        for level in available_levels:
            sheet_name = f"{zone} {LEVEL_LABELS.get(level, level)}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = ZONE_TAB_COLORS.get(zone, "888888")

            _write_level_sheet(
                ws, zone, level, zone_data, profiles,
                available_profiles,
            )

    _write_summary(wb, data)
    wb.save(output_path)


def _write_level_sheet(
    ws, zone: str, level: str, zone_data: dict,
    profiles: dict, available_profiles: list[str],
) -> None:
    """Write one granularity level for one zone."""
    time_cols = LEVEL_TIME_COLS[level]
    is_hourly = level == "hourly"

    # Build header
    headers = [label for _, label in time_cols]
    headers.append("Baseload")

    col_map: list[tuple[str, str]] = []
    for pk in available_profiles:
        name = profiles.get(pk, pk)
        headers.append(name)
        col_map.append((pk, "capture"))
        if is_hourly:
            headers.append("Vikt")
            col_map.append((pk, "weight"))
        else:
            headers.append("Ratio")
            col_map.append((pk, "ratio"))

    # Write header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Baseload rows define ordering
    baseload_rows = zone_data.get("baseload", {}).get(level, [])
    if not baseload_rows:
        return

    # Build lookup per profile
    profile_lookups: dict[str, dict] = {}
    for pk in available_profiles:
        rows = zone_data.get(pk, {}).get(level, [])
        lookup = {}
        for r in rows:
            lookup[_level_key(level, r)] = r
        profile_lookups[pk] = lookup

    # Write data rows
    for row_idx, bl_row in enumerate(baseload_rows, 2):
        col = 1

        for field, _ in time_cols:
            val = bl_row.get(field)
            if field == "month" and isinstance(val, int) and 1 <= val <= 12:
                val = MONTH_NAMES[val - 1]
            ws.cell(row=row_idx, column=col, value=val).font = BODY_FONT
            col += 1

        _write_price_cell(ws, row_idx, col, bl_row.get("baseload"))
        col += 1

        lk = _level_key(level, bl_row)
        for pk, col_type in col_map:
            p_row = profile_lookups.get(pk, {}).get(lk)
            val = p_row.get(col_type) if p_row else None
            if col_type == "capture":
                _write_price_cell(ws, row_idx, col, val)
            elif col_type == "ratio":
                _write_ratio_cell(ws, row_idx, col, val)
            elif col_type == "weight":
                _write_weight_cell(ws, row_idx, col, val)
            col += 1

    # Freeze + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(baseload_rows) + 1}"
    )

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _write_summary(wb: Workbook, data: dict) -> None:
    """Write summary sheet with latest year per zone."""
    ws = wb.create_sheet(title="Sammanfattning", index=0)

    profiles = data.get("profiles", {})
    profile_meta = data.get("profile_meta", {})
    all_data = data.get("data", {})
    zones = data.get("zones", [])

    all_available: list[str] = []
    for zone in zones:
        zd = all_data.get(zone, {})
        for k in profiles:
            if (k != "baseload" and k in zd and k not in all_available
                    and not profile_meta.get(k, {}).get("category", "").startswith("bess")):
                all_available.append(k)

    headers = ["Zon", "År", "Baseload"]
    for pk in all_available:
        headers.append(profiles.get(pk, pk))
        headers.append("Ratio")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for zone in zones:
        zd = all_data.get(zone, {})
        yearly = zd.get("baseload", {}).get("yearly", [])
        if not yearly:
            continue

        latest = yearly[-1]
        col = 1
        ws.cell(row=row, column=col, value=zone).font = Font(
            name="Calibri", bold=True, size=11,
        )
        col += 1
        ws.cell(row=row, column=col, value=latest["year"]).font = BODY_FONT
        col += 1
        _write_price_cell(ws, row, col, latest.get("baseload"))
        col += 1

        for pk in all_available:
            pk_yearly = zd.get(pk, {}).get("yearly", [])
            match = next(
                (r for r in pk_yearly if r["year"] == latest["year"]), None,
            )
            _write_price_cell(ws, row, col, match["capture"] if match else None)
            col += 1
            _write_ratio_cell(ws, row, col, match["ratio"] if match else None)
            col += 1

        row += 1

    ws.freeze_panes = "A2"
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _level_key(level: str, row: dict) -> tuple:
    """Create a lookup key for a row at a given level."""
    if level == "yearly":
        return (row.get("year"),)
    if level == "monthly":
        return (row.get("year"), row.get("month"))
    if level == "daily":
        return (row.get("date"),)
    if level == "hourly":
        return (row.get("date"), row.get("hour"))
    return ()


def _write_price_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
        cell.number_format = NUM_FMT_PRICE
    else:
        cell.value = "–"
    cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="right")


def _write_ratio_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
        cell.number_format = NUM_FMT_RATIO
        if value >= 0.9:
            cell.font = Font(name="Calibri", size=10, color="059669")
        elif value >= 0.7:
            cell.font = Font(name="Calibri", size=10, color="D97706")
        else:
            cell.font = Font(name="Calibri", size=10, color="DC2626")
    else:
        cell.value = "–"
        cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="right")


def _write_weight_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
        cell.number_format = NUM_FMT_WEIGHT
    else:
        cell.value = "–"
    cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="right")
