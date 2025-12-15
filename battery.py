#!/usr/bin/env python3
"""Battery arbitrage analysis CLI."""

from __future__ import annotations

import argparse
from datetime import datetime, date
from pathlib import Path

from elpris.battery import (
    ROUND_TRIP_EFFICIENCY,
    REPORTS_DIR,
    extract_daily_stats,
    extract_hourly_profile,
    calculate_1cycle_revenue,
    calculate_2cycle_revenue,
    aggregate_by_month,
    aggregate_by_year,
    format_terminal_table,
    # Optimal DP-based functions
    calculate_optimal_1cycle_revenue,
    calculate_optimal_2cycle_revenue,
    calculate_optimal_daily_arbitrage,
    format_optimal_schedule,
)
from elpris.config import ZONES
from elpris.battery_excel import export_battery_excel


def export_raw_excel(zones: list[str], filename: str | None = None) -> Path:
    """Export raw daily and hourly data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"battery_raw_data_{timestamp}.xlsx"

    output_path = REPORTS_DIR / filename
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Sheet 1: Daily stats per zone
    for zone in zones:
        ws = wb.create_sheet(f"Daily {zone}")

        headers = ["Date", "Min EUR/MWh", "Max EUR/MWh", "Spread", "Avg EUR/MWh", "Median EUR/MWh", "Std EUR/MWh", "Min Hour", "Max Hour", "Records"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        daily = extract_daily_stats(zone)
        for row, d in enumerate(daily, 2):
            ws.cell(row=row, column=1, value=d["date"].isoformat())
            ws.cell(row=row, column=2, value=d["min_eur"])
            ws.cell(row=row, column=3, value=d["max_eur"])
            ws.cell(row=row, column=4, value=d["spread"])
            ws.cell(row=row, column=5, value=d["avg_eur"])
            ws.cell(row=row, column=6, value=d.get("median_eur", 0))
            ws.cell(row=row, column=7, value=d["std_eur"])
            ws.cell(row=row, column=8, value=d["min_hour"])
            ws.cell(row=row, column=9, value=d["max_hour"])
            ws.cell(row=row, column=10, value=d["records"])

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "A2"

    # Sheet 2: Hourly profile (all zones)
    ws = wb.create_sheet("Hourly Profile")
    headers = ["Hour"] + [f"{z} Avg EUR/MWh" for z in zones]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Get hourly profiles for all zones
    hourly_data = {zone: extract_hourly_profile(zone) for zone in zones}

    for hour in range(24):
        row = hour + 2
        ws.cell(row=row, column=1, value=f"{hour:02d}:00")
        for col, zone in enumerate(zones, 2):
            price = hourly_data[zone].get(hour, 0)
            ws.cell(row=row, column=col, value=round(price, 2))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def export_analysis_excel(
    zones: list[str],
    efficiency: float = ROUND_TRIP_EFFICIENCY,
    filename: str | None = None,
) -> Path:
    """Export full analysis to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"battery_analysis_{timestamp}.xlsx"

    output_path = REPORTS_DIR / filename
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    bold = Font(bold=True)

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    headers = ["Zone", "Year", "Avg Spread", "Median Spread", "Total Revenue 1C", "Total Revenue 2C", "Profitable Days %"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    summary_row = 2
    for zone in zones:
        daily = extract_daily_stats(zone)
        if not daily:
            continue

        yearly = aggregate_by_year(daily)
        rev_1c = calculate_1cycle_revenue(daily, efficiency)
        rev_2c = calculate_2cycle_revenue(zone, efficiency=efficiency)

        for y in yearly:
            year = y["period"]
            # Filter daily revenues for this year
            year_daily = [d for d in daily if d["date"].year == int(year)]
            year_rev_1c = calculate_1cycle_revenue(year_daily, efficiency)
            year_rev_2c = calculate_2cycle_revenue(zone, year=int(year), efficiency=efficiency)

            ws_summary.cell(row=summary_row, column=1, value=zone)
            ws_summary.cell(row=summary_row, column=2, value=year)
            ws_summary.cell(row=summary_row, column=3, value=y["avg_spread"])
            ws_summary.cell(row=summary_row, column=4, value=y.get("median_spread", 0))
            ws_summary.cell(row=summary_row, column=5, value=year_rev_1c["total_revenue_eur"])
            ws_summary.cell(row=summary_row, column=6, value=year_rev_2c["total_revenue_eur"])
            ws_summary.cell(row=summary_row, column=7, value=year_rev_1c["profitable_pct"])
            summary_row += 1

    for col in range(1, len(headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    ws_summary.freeze_panes = "A2"

    # Per-zone detailed sheets
    for zone in zones:
        daily = extract_daily_stats(zone)
        if not daily:
            continue

        monthly = aggregate_by_month(daily)
        rev_1c = calculate_1cycle_revenue(daily, efficiency)

        ws = wb.create_sheet(zone)
        headers = ["Period", "Avg Spread", "Median Spread", "Max Spread", "Days", "Revenue 1C", "Revenue 2C"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row = 2
        current_year = None

        for m in monthly:
            period = m["period"]
            year = period.split("-")[0]

            # Year summary row when year changes
            if current_year and year != current_year:
                # Calculate year totals
                from statistics import median as calc_median
                year_months = [x for x in monthly if x["period"].startswith(current_year)]
                year_daily = [d for d in daily if d["date"].strftime("%Y") == current_year]
                year_spreads = [d["spread"] for d in year_daily]
                year_median = calc_median(year_spreads) if year_spreads else 0
                year_rev_1c = calculate_1cycle_revenue(year_daily, efficiency)
                year_rev_2c = calculate_2cycle_revenue(zone, year=int(current_year), efficiency=efficiency)

                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = summary_fill
                    ws.cell(row=row, column=col).font = bold

                ws.cell(row=row, column=1, value=current_year)
                ws.cell(row=row, column=2, value=round(sum(x["avg_spread"] for x in year_months) / len(year_months), 2))
                ws.cell(row=row, column=3, value=round(year_median, 2))
                ws.cell(row=row, column=4, value=max(x["max_spread"] for x in year_months))
                ws.cell(row=row, column=5, value=sum(x["days"] for x in year_months))
                ws.cell(row=row, column=6, value=year_rev_1c["total_revenue_eur"])
                ws.cell(row=row, column=7, value=year_rev_2c["total_revenue_eur"])
                row += 1

            current_year = year

            # Find monthly revenues
            month_daily = [d for d in daily if d["date"].strftime("%Y-%m") == period]
            month_rev_1c = calculate_1cycle_revenue(month_daily, efficiency)
            month_rev_2c_result = calculate_2cycle_revenue(zone, efficiency=efficiency)
            # Filter 2c revenues for this month
            month_rev_2c = sum(
                r["total_revenue"] for r in month_rev_2c_result["daily_revenues"]
                if r["date"].strftime("%Y-%m") == period
            )

            ws.cell(row=row, column=1, value=period)
            ws.cell(row=row, column=2, value=m["avg_spread"])
            ws.cell(row=row, column=3, value=m.get("median_spread", 0))
            ws.cell(row=row, column=4, value=m["max_spread"])
            ws.cell(row=row, column=5, value=m["days"])
            ws.cell(row=row, column=6, value=month_rev_1c["total_revenue_eur"])
            ws.cell(row=row, column=7, value=round(month_rev_2c, 2))
            row += 1

        # Final year summary
        if current_year:
            from statistics import median as calc_median
            year_months = [x for x in monthly if x["period"].startswith(current_year)]
            year_daily = [d for d in daily if d["date"].strftime("%Y") == current_year]
            year_spreads = [d["spread"] for d in year_daily]
            year_median = calc_median(year_spreads) if year_spreads else 0
            year_rev_1c = calculate_1cycle_revenue(year_daily, efficiency)
            year_rev_2c = calculate_2cycle_revenue(zone, year=int(current_year), efficiency=efficiency)

            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = summary_fill
                ws.cell(row=row, column=col).font = bold

            ws.cell(row=row, column=1, value=current_year)
            ws.cell(row=row, column=2, value=round(sum(x["avg_spread"] for x in year_months) / len(year_months), 2))
            ws.cell(row=row, column=3, value=round(year_median, 2))
            ws.cell(row=row, column=4, value=max(x["max_spread"] for x in year_months))
            ws.cell(row=row, column=5, value=sum(x["days"] for x in year_months))
            ws.cell(row=row, column=6, value=year_rev_1c["total_revenue_eur"])
            ws.cell(row=row, column=7, value=year_rev_2c["total_revenue_eur"])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def print_analysis(
    zone: str,
    analysis_type: str = "1cycle",
    efficiency: float = ROUND_TRIP_EFFICIENCY,
    year: int | None = None,
):
    """Print arbitrage analysis to terminal."""
    daily = extract_daily_stats(zone, year)

    if not daily:
        print(f"No data available for {zone}" + (f" {year}" if year else ""))
        return

    monthly = aggregate_by_month(daily)

    if analysis_type == "1cycle":
        revenue = calculate_1cycle_revenue(daily, efficiency)
    else:
        revenue = calculate_2cycle_revenue(zone, year, efficiency)

    table = format_terminal_table(monthly, revenue, zone)
    print(table)


def main():
    parser = argparse.ArgumentParser(
        description="Battery arbitrage analysis for Swedish electricity zones"
    )
    parser.add_argument(
        "--zone",
        choices=ZONES,
        default="SE3",
        help="Electricity zone (default: SE3)",
    )
    parser.add_argument(
        "--zones",
        nargs="+",
        choices=ZONES,
        default=None,
        help="Multiple zones for export (default: all SE1-SE4)",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Filter to specific year",
    )
    parser.add_argument(
        "--analysis",
        choices=["1cycle", "2cycle"],
        default=None,
        help="Run arbitrage analysis (1cycle or 2cycle)",
    )
    parser.add_argument(
        "--efficiency",
        type=float,
        default=ROUND_TRIP_EFFICIENCY,
        help=f"Round-trip efficiency (default: {ROUND_TRIP_EFFICIENCY})",
    )
    parser.add_argument(
        "--export-raw",
        action="store_true",
        help="Export raw daily/hourly data to Excel",
    )
    parser.add_argument(
        "--excel",
        type=str,
        nargs="?",
        const="",
        default=None,
        help="Export full analysis to Excel (optional filename)",
    )
    parser.add_argument(
        "--optimal",
        action="store_true",
        help="Use optimal DP algorithm (recommended, accounts for SoC constraints)",
    )
    parser.add_argument(
        "--day",
        type=date.fromisoformat,
        default=None,
        help="Show optimal schedule for specific day (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--cycles",
        type=int,
        choices=[1, 2],
        default=1,
        help="Number of cycles for optimal analysis (default: 1)",
    )

    args = parser.parse_args()

    zones = args.zones if args.zones else list(ZONES)

    if args.export_raw:
        print(f"Exporting raw data for zones: {', '.join(zones)}")
        output_path = export_raw_excel(zones)
        print(f"Exported: {output_path}")
        return

    if args.excel is not None:
        filename = args.excel if args.excel else None
        print(f"Exporting battery arbitrage report for zones: {', '.join(zones)}")
        print(f"Efficiency: {args.efficiency*100:.0f}%")
        output_path = export_battery_excel(filename, args.efficiency, zones)
        print(f"Exported: {output_path}")
        return

    # Show optimal schedule for a specific day
    if args.day:
        print(f"Zone: {args.zone}")
        print(f"Day: {args.day}")
        print(f"Cycles: {args.cycles}")
        print(f"Efficiency: {args.efficiency*100:.0f}%")

        result = calculate_optimal_daily_arbitrage(
            args.zone, args.day, args.cycles, args.efficiency
        )
        print(format_optimal_schedule(result, args.day))
        return

    # Optimal analysis for a zone/year
    if args.optimal:
        print(f"Zone: {args.zone}")
        print(f"Cycles: {args.cycles}")
        print(f"Efficiency: {args.efficiency*100:.0f}%")
        if args.year:
            print(f"Year: {args.year}")
        print()
        print("Running optimal DP analysis (this may take a moment)...")
        print()

        if args.cycles == 1:
            result = calculate_optimal_1cycle_revenue(args.zone, args.year, args.efficiency)
        else:
            result = calculate_optimal_2cycle_revenue(args.zone, args.year, args.efficiency)

        print(f"Optimal {args.cycles}-Cycle Battery Arbitrage - {args.zone}")
        print("=" * 70)
        print(f"Total Revenue:     {result['total_revenue_eur']:>12,.2f} EUR/MWh")
        print(f"Avg Daily Revenue: {result['avg_daily_revenue']:>12,.2f} EUR/MWh")
        print(f"Profitable Days:   {result['profitable_days']:>12} / {result['total_days']}")
        print(f"Profitable %:      {result['profitable_pct']:>12.1f}%")
        return

    if args.analysis:
        print(f"Zone: {args.zone}")
        print(f"Analysis: {args.analysis} (legacy)")
        print(f"Efficiency: {args.efficiency*100:.0f}%")
        if args.year:
            print(f"Year: {args.year}")
        print()
        print_analysis(args.zone, args.analysis, args.efficiency, args.year)
        return

    # Default: show 1-cycle analysis for specified zone (legacy)
    print(f"Zone: {args.zone}")
    print(f"Efficiency: {args.efficiency*100:.0f}%")
    if args.year:
        print(f"Year: {args.year}")
    print()
    print("Note: Using legacy analysis. Add --optimal for accurate SoC-constrained results.")
    print()
    print_analysis(args.zone, "1cycle", args.efficiency, args.year)


if __name__ == "__main__":
    main()
