"""Baseload PPA analysis with solar + wind + battery.

Analyzes how to deliver 100% constant baseload by combining:
- Solar production (PVsyst profile)
- Wind production (ENTSO-E actual data)
- Battery storage to fill production gaps
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from .config import DATA_DIR
from .entsoe_profile import load_entsoe_generation, ENTSOE_DIR


@dataclass
class BaseloadResult:
    """Results from baseload analysis."""
    sol_mw: float
    vind_mw: float
    baseload_mw: float
    total_production_mwh: float
    total_hours: int
    deficit_hours: int
    surplus_hours: int
    max_deficit_mw: float
    max_surplus_mw: float
    battery_mwh_required: float
    battery_mw_required: float
    sol_capacity_factor: float
    vind_capacity_factor: float
    combined_capacity_factor: float
    efficiency: float = 0.90  # Round-trip battery efficiency


@dataclass
class RatioGridResult:
    """Result from a single solar/wind ratio analysis."""
    sol_mw: float
    vind_mw: float
    total_mw: float
    sol_ratio: float              # sol_mw / total_mw (0.0 - 1.0)
    mean_production_mw: float     # Average hourly production
    baseload_mw: float            # Target baseload for this analysis
    baseload_pct: float           # baseload_mw / mean_production
    battery_mwh: float
    battery_mw: float
    battery_duration_h: float     # battery_mwh / battery_mw
    meets_constraint: bool        # True if duration in target range
    sol_cf: float
    vind_cf: float
    combined_cf: float
    deficit_hours: int
    surplus_hours: int


@dataclass
class RatioOptimizationResult:
    """Complete optimization result with grid search data."""
    zone: str
    year: int
    sol_profile_name: str
    total_capacity_mw: float
    battery_duration_min: float
    battery_duration_max: float
    efficiency: float  # Round-trip battery efficiency
    all_results: list
    optimal: Optional[RatioGridResult]
    best_per_constraint: list  # Best results that meet constraint


def load_pvsyst_profile(name: str) -> dict[tuple[int, int, int], float]:
    """
    Load PVsyst solar profile.

    Returns dict mapping (month, day, hour) -> power_mw (per MW installed)
    """
    filepath = DATA_DIR / "solar_profiles" / f"{name}.csv"

    if not filepath.exists():
        raise FileNotFoundError(f"PVsyst profile not found: {filepath}")

    profile = {}
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (int(row["month"]), int(row["day"]), int(row["hour"]))
            profile[key] = float(row["power_mw"])

    return profile


def load_wind_profile_normalized(zone: str, year: int) -> dict[tuple[int, int, int], float]:
    """
    Load ENTSO-E wind data and normalize to capacity factor.

    Since installed capacity changes over time, we normalize each year's data
    by dividing by the max observed production (approximates installed capacity).

    Returns dict mapping (month, day, hour) -> capacity_factor (0-1)
    """
    records = load_entsoe_generation(zone, "wind_onshore", [year])

    if not records:
        return {}

    # Parse into dict
    hourly_data: dict[tuple[int, int, int], float] = {}

    for rec in records:
        ts = datetime.fromisoformat(rec["time_start"].replace("Z", "+00:00"))
        # Convert UTC to local time (CET = UTC+1, simplified)
        local_hour = (ts.hour + 1) % 24
        key = (ts.month, ts.day, local_hour)
        hourly_data[key] = rec["generation_mw"]

    # Normalize by max production (approximates capacity factor)
    max_production = max(hourly_data.values()) if hourly_data else 1.0

    normalized = {k: v / max_production for k, v in hourly_data.items()}

    return normalized


def load_wind_profile_from_typical(zone: str) -> dict[tuple[int, int, int], float]:
    """
    Load pre-computed typical wind profile.

    These are already normalized (sum = 1.0 over year).
    We need to convert to capacity factor style.
    """
    filepath = DATA_DIR / "profiles" / f"wind_onshore_{zone}.csv"

    if not filepath.exists():
        raise FileNotFoundError(f"Wind profile not found: {filepath}")

    profile = {}
    total_weight = 0.0

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (int(row["month"]), int(row["day"]), int(row["hour"]))
            weight = float(row["weight"])
            profile[key] = weight
            total_weight += weight

    # Convert from normalized weights to approximate capacity factor
    # Typical wind CF in Sweden is ~30%, so scale accordingly
    # 8760 hours * 0.30 CF = 2628 full load hours
    # If sum of weights = 1.0, each weight represents fraction of annual production
    # To get hourly CF: weight * 8760 * CF_annual

    # Actually, we need to think about this differently:
    # The weight represents relative production at that hour
    # To convert to CF-like values, multiply by hours in year * assumed CF

    assumed_annual_cf = 0.30  # 30% typical for wind in SE3
    hours_per_year = len(profile)  # Should be ~8760

    # Scale so mean value equals assumed CF
    mean_weight = total_weight / len(profile)
    scale_factor = assumed_annual_cf / mean_weight if mean_weight > 0 else 1.0

    cf_profile = {k: v * scale_factor for k, v in profile.items()}

    return cf_profile


def generate_hourly_timeseries(
    year: int,
    sol_profile: dict[tuple[int, int, int], float],
    vind_profile: dict[tuple[int, int, int], float],
    sol_mw: float,
    vind_mw: float,
) -> list[dict]:
    """
    Generate hourly production timeseries for a full year.

    Returns list of dicts with timestamp, sol_mw, vind_mw, total_mw
    """
    from datetime import timedelta

    # Start from Jan 1
    start = datetime(year, 1, 1, 0, 0)

    timeseries = []
    current = start

    while current.year == year:
        key = (current.month, current.day, current.hour)

        sol_output = sol_profile.get(key, 0.0) * sol_mw
        vind_output = vind_profile.get(key, 0.0) * vind_mw

        timeseries.append({
            "timestamp": current,
            "sol_mw": sol_output,
            "vind_mw": vind_output,
            "total_mw": sol_output + vind_output,
        })

        current += timedelta(hours=1)

    return timeseries


def calculate_battery_requirement(
    timeseries: list[dict],
    baseload_mw: float,
    efficiency: float = 0.90,
) -> dict:
    """
    Simulate battery operation to determine required size.

    The battery must be large enough to:
    1. Store surplus during high production
    2. Deliver deficit during low production

    Uses the "cumulative sum" method: track cumulative energy balance
    over the year. Battery size = max cumulative deficit before recovery.

    Args:
        timeseries: Hourly production data
        baseload_mw: Target baseload power level
        efficiency: Round-trip battery efficiency (default 0.90 = 90%)
                   Applied during charging: stored = surplus × efficiency

    Returns dict with battery sizing info.
    """
    # Calculate deficit and surplus for each hour
    deficits = []
    surpluses = []
    net_flows = []  # positive = charging, negative = discharging

    for row in timeseries:
        diff = row["total_mw"] - baseload_mw
        if diff < 0:
            # Deficit: need to discharge (no efficiency loss on discharge)
            deficits.append(abs(diff))
            surpluses.append(0)
            net_flows.append(diff)  # negative, full amount needed
        else:
            # Surplus: charge battery (efficiency loss during charging)
            deficits.append(0)
            surpluses.append(diff)
            net_flows.append(diff * efficiency)  # positive, reduced by efficiency

    # Method 1: Find max drawdown in cumulative sum
    # This finds the largest "valley" in the cumulative energy balance
    cumsum = 0.0
    max_cumsum = 0.0
    max_drawdown = 0.0

    for flow in net_flows:
        cumsum += flow
        max_cumsum = max(max_cumsum, cumsum)
        drawdown = max_cumsum - cumsum
        max_drawdown = max(max_drawdown, drawdown)

    # Method 2: Simulate with wraparound (year is cyclic)
    # Run simulation twice to handle year boundary
    extended_flows = net_flows + net_flows
    battery_level = 0.0
    min_level = 0.0

    for flow in extended_flows:
        battery_level += flow
        min_level = min(min_level, battery_level)

    # Battery size is the larger of the two methods
    battery_mwh_required = max(max_drawdown, abs(min_level))

    # Also need to check if we can even achieve balance
    # (total surplus × efficiency must >= total deficit for feasibility)
    total_surplus = sum(surpluses)
    total_deficit = sum(deficits)
    effective_surplus = total_surplus * efficiency

    # If total deficit > effective surplus, it's impossible
    if total_deficit > effective_surplus + 0.001:
        # Need infinite battery (or can't achieve this baseload)
        battery_mwh_required = float('inf')

    # Power requirement is max hourly deficit
    battery_mw_required = max(deficits) if deficits else 0

    return {
        "battery_mwh_required": battery_mwh_required,
        "battery_mw_required": battery_mw_required,
        "total_surplus_mwh": total_surplus,
        "total_deficit_mwh": total_deficit,
        "effective_surplus_mwh": effective_surplus,
        "balance_mwh": effective_surplus - total_deficit,
        "efficiency": efficiency,
        "deficit_hours": sum(1 for d in deficits if d > 0),
        "surplus_hours": sum(1 for s in surpluses if s > 0),
        "max_deficit_mw": max(deficits) if deficits else 0,
        "max_surplus_mw": max(surpluses) if surpluses else 0,
    }


def run_baseload_analysis(
    sol_mw: float = 1.0,
    vind_mw: float = 1.0,
    zone: str = "SE3",
    sol_profile_name: str = "south_lundby",
    year: int = 2024,
    baseload_mw: Optional[float] = None,
    efficiency: float = 0.90,
) -> BaseloadResult:
    """
    Run complete baseload analysis.

    Args:
        sol_mw: Installed solar capacity (MW)
        vind_mw: Installed wind capacity (MW)
        zone: Electricity zone (SE1-SE4)
        sol_profile_name: PVsyst profile name
        year: Year to analyze
        baseload_mw: Target baseload (None = use mean production)
        efficiency: Round-trip battery efficiency (default 0.90 = 90%)

    Returns:
        BaseloadResult with all analysis data
    """
    # Load profiles
    sol_profile = load_pvsyst_profile(sol_profile_name)
    vind_profile = load_wind_profile_from_typical(zone)

    # Generate hourly timeseries
    timeseries = generate_hourly_timeseries(
        year, sol_profile, vind_profile, sol_mw, vind_mw
    )

    # Calculate capacity factors
    total_hours = len(timeseries)
    sol_production = sum(row["sol_mw"] for row in timeseries)
    vind_production = sum(row["vind_mw"] for row in timeseries)
    total_production = sum(row["total_mw"] for row in timeseries)

    sol_cf = sol_production / (sol_mw * total_hours) if sol_mw > 0 else 0
    vind_cf = vind_production / (vind_mw * total_hours) if vind_mw > 0 else 0
    combined_cf = total_production / ((sol_mw + vind_mw) * total_hours)

    # Determine baseload level
    mean_production = total_production / total_hours
    if baseload_mw is None:
        baseload_mw = mean_production

    # Calculate battery requirement with efficiency
    battery_info = calculate_battery_requirement(timeseries, baseload_mw, efficiency)

    return BaseloadResult(
        sol_mw=sol_mw,
        vind_mw=vind_mw,
        baseload_mw=baseload_mw,
        total_production_mwh=total_production,
        total_hours=total_hours,
        deficit_hours=battery_info["deficit_hours"],
        surplus_hours=battery_info["surplus_hours"],
        max_deficit_mw=battery_info["max_deficit_mw"],
        max_surplus_mw=battery_info["max_surplus_mw"],
        battery_mwh_required=battery_info["battery_mwh_required"],
        battery_mw_required=battery_info["battery_mw_required"],
        sol_capacity_factor=sol_cf,
        vind_capacity_factor=vind_cf,
        combined_capacity_factor=combined_cf,
        efficiency=efficiency,
    )


def analyze_wind_ratios(
    sol_mw: float = 1.0,
    zone: str = "SE3",
    sol_profile_name: str = "south_lundby",
    year: int = 2024,
    wind_ratios: Optional[list[float]] = None,
) -> list[BaseloadResult]:
    """
    Analyze different wind capacities relative to solar.

    Args:
        sol_mw: Fixed solar capacity
        zone: Electricity zone
        sol_profile_name: PVsyst profile name
        year: Year to analyze
        wind_ratios: List of wind MW values to test (default: 0.5, 1, 2, 3, 5)

    Returns:
        List of BaseloadResult for each wind ratio
    """
    if wind_ratios is None:
        wind_ratios = [0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 5.0]

    results = []
    for vind_mw in wind_ratios:
        result = run_baseload_analysis(
            sol_mw=sol_mw,
            vind_mw=vind_mw,
            zone=zone,
            sol_profile_name=sol_profile_name,
            year=year,
        )
        results.append(result)

    return results


def print_analysis_table(results: list[BaseloadResult]):
    """Print results as a formatted table."""
    print("\n" + "=" * 90)
    print("BASELOAD ANALYSIS: Sol + Vind + Batteri")
    print("=" * 90)

    if results:
        r0 = results[0]
        print(f"Sol: {r0.sol_mw} MW | Zon: SE3 | Sol CF: {r0.sol_capacity_factor:.1%}")
        print("-" * 90)

    print(f"{'Vind':<8} {'Vind CF':<10} {'Baseload':<10} {'Deficit':<12} {'Batteri':<12} {'Batteri':<10}")
    print(f"{'(MW)':<8} {'(%)':<10} {'(MW)':<10} {'(timmar)':<12} {'(MWh)':<12} {'(MW)':<10}")
    print("-" * 90)

    for r in results:
        print(
            f"{r.vind_mw:<8.1f} "
            f"{r.vind_capacity_factor:<10.1%} "
            f"{r.baseload_mw:<10.3f} "
            f"{r.deficit_hours:<12} "
            f"{r.battery_mwh_required:<12.1f} "
            f"{r.battery_mw_required:<10.3f}"
        )

    print("-" * 90)
    print("\nOBS: Batteri (MWh) är minsta kapacitet för 100% leverans.")
    print("     Batteri (MW) är max effektbehov vid underskott.")


def find_max_baseload_for_duration(
    timeseries: list[dict],
    duration_min: float,
    duration_max: float,
    mean_production: float,
    n_samples: int = 100,
    efficiency: float = 0.90,
) -> tuple[float, float, float, int, int]:
    """
    Find the maximum baseload that results in battery duration within target range.

    Uses grid search over baseload levels to find solutions, as the relationship
    between baseload and battery duration is often non-monotonic.

    Args:
        timeseries: Hourly production data
        duration_min: Minimum acceptable battery duration (hours)
        duration_max: Maximum acceptable battery duration (hours)
        mean_production: Mean hourly production (upper bound for baseload)
        n_samples: Number of baseload levels to test
        efficiency: Round-trip battery efficiency (default 0.90 = 90%)

    Returns:
        Tuple of (baseload_mw, battery_mwh, battery_mw, deficit_hours, surplus_hours)
    """
    best_result = None
    best_baseload = 0.0
    best_distance = float('inf')

    # Test baseload levels from 5% to 100% of mean production
    for i in range(1, n_samples + 1):
        baseload = mean_production * (i / n_samples)

        battery_info = calculate_battery_requirement(timeseries, baseload, efficiency)
        battery_mwh = battery_info["battery_mwh_required"]
        battery_mw = battery_info["battery_mw_required"]

        # Skip infeasible solutions
        if battery_mwh == float('inf'):
            continue

        # Calculate duration based on baseload (correct definition for PPA)
        # Duration = how long battery can cover FULL baseload at zero production
        if baseload > 0.001:  # Minimum threshold
            duration = battery_mwh / baseload
        else:
            duration = 0.0

        # Check if within constraint
        if duration_min <= duration <= duration_max:
            # Valid solution - keep the highest baseload
            if baseload > best_baseload:
                best_result = battery_info
                best_baseload = baseload
        else:
            # Track closest to constraint for fallback
            if duration < duration_min:
                distance = duration_min - duration
            else:
                distance = duration - duration_max

            if distance < best_distance and baseload > 0.01:
                best_distance = distance
                if best_result is None:  # Only use as fallback
                    best_result = battery_info
                    best_baseload = baseload

    # If no solution found, return last attempt
    if best_result is None:
        battery_info = calculate_battery_requirement(timeseries, mean_production * 0.5, efficiency)
        return (
            mean_production * 0.5,
            battery_info["battery_mwh_required"],
            battery_info["battery_mw_required"],
            battery_info["deficit_hours"],
            battery_info["surplus_hours"],
        )

    return (
        best_baseload,
        best_result["battery_mwh_required"],
        best_result["battery_mw_required"],
        best_result["deficit_hours"],
        best_result["surplus_hours"],
    )


def optimize_solar_wind_ratio(
    total_capacity_mw: float = 2.0,
    zone: str = "SE3",
    sol_profile_name: str = "south_lundby",
    year: int = 2024,
    n_steps: int = 21,
    battery_duration_min: float = 1.0,
    battery_duration_max: float = 3.0,
    efficiency: float = 0.90,
) -> RatioOptimizationResult:
    """
    Find optimal solar/wind ratio for baseload PPA with battery constraints.

    Keeps total capacity constant and varies the split between solar and wind.
    For each ratio, finds the maximum baseload achievable within the battery
    duration constraint.

    Args:
        total_capacity_mw: Total installed capacity (solar + wind)
        zone: Electricity zone (SE1-SE4)
        sol_profile_name: PVsyst profile name
        year: Analysis year
        n_steps: Number of ratio steps (21 = 5% increments)
        battery_duration_min: Minimum acceptable battery duration (hours)
        battery_duration_max: Maximum acceptable battery duration (hours)
        efficiency: Round-trip battery efficiency (default 0.90 = 90%)

    Returns:
        RatioOptimizationResult with all grid points and identified optimum
    """
    # Load profiles once
    sol_profile = load_pvsyst_profile(sol_profile_name)
    vind_profile = load_wind_profile_from_typical(zone)

    all_results = []

    for i in range(n_steps):
        sol_fraction = i / (n_steps - 1)
        sol_mw = total_capacity_mw * sol_fraction
        vind_mw = total_capacity_mw * (1 - sol_fraction)

        # Generate timeseries for this ratio
        timeseries = generate_hourly_timeseries(
            year, sol_profile, vind_profile, sol_mw, vind_mw
        )

        # Calculate capacity factors
        total_hours = len(timeseries)
        sol_production = sum(row["sol_mw"] for row in timeseries)
        vind_production = sum(row["vind_mw"] for row in timeseries)
        total_production = sum(row["total_mw"] for row in timeseries)

        sol_cf = sol_production / (sol_mw * total_hours) if sol_mw > 0 else 0
        vind_cf = vind_production / (vind_mw * total_hours) if vind_mw > 0 else 0
        combined_cf = total_production / (total_capacity_mw * total_hours) if total_capacity_mw > 0 else 0

        mean_production = total_production / total_hours

        # Find max baseload within duration constraint
        baseload_mw, battery_mwh, battery_mw, deficit_hours, surplus_hours = \
            find_max_baseload_for_duration(
                timeseries,
                battery_duration_min,
                battery_duration_max,
                mean_production,
                efficiency=efficiency,
            )

        # Calculate duration based on baseload (not battery_mw)
        # Duration = how long battery can cover FULL baseload at zero production
        # This is the correct definition for baseload PPA guarantee
        if baseload_mw > 0 and battery_mwh < float('inf'):
            duration = battery_mwh / baseload_mw
        else:
            duration = float('inf')

        # Check if meets constraint
        meets_constraint = (
            battery_duration_min <= duration <= battery_duration_max
            and battery_mwh < float('inf')
        )

        # Calculate baseload as percentage of mean production
        baseload_pct = baseload_mw / mean_production if mean_production > 0 else 0

        result = RatioGridResult(
            sol_mw=sol_mw,
            vind_mw=vind_mw,
            total_mw=total_capacity_mw,
            sol_ratio=sol_fraction,
            mean_production_mw=mean_production,
            baseload_mw=baseload_mw,
            baseload_pct=baseload_pct,
            battery_mwh=battery_mwh,
            battery_mw=battery_mw,
            battery_duration_h=duration,
            meets_constraint=meets_constraint,
            sol_cf=sol_cf,
            vind_cf=vind_cf,
            combined_cf=combined_cf,
            deficit_hours=deficit_hours,
            surplus_hours=surplus_hours,
        )
        all_results.append(result)

    # Find optimal: highest baseload_pct that meets constraint
    valid_results = [r for r in all_results if r.meets_constraint]
    if valid_results:
        optimal = max(valid_results, key=lambda r: r.baseload_pct)
    else:
        # No valid result - find closest to constraint
        optimal = min(all_results, key=lambda r: abs(r.battery_duration_h - battery_duration_max))

    return RatioOptimizationResult(
        zone=zone,
        year=year,
        sol_profile_name=sol_profile_name,
        total_capacity_mw=total_capacity_mw,
        battery_duration_min=battery_duration_min,
        battery_duration_max=battery_duration_max,
        efficiency=efficiency,
        all_results=all_results,
        optimal=optimal,
        best_per_constraint=valid_results,
    )


def print_ratio_optimization_table(result: RatioOptimizationResult) -> None:
    """Print optimization results as formatted table."""
    print("\n" + "=" * 100)
    print("OPTIMAL SOL/VIND RATIO - Baseload PPA med batteri-constraint")
    print("=" * 100)
    print(f"Zon: {result.zone} | År: {result.year} | Profil: {result.sol_profile_name}")
    print(f"Total kapacitet: {result.total_capacity_mw} MW | Batteri-duration: {result.battery_duration_min}-{result.battery_duration_max}h | Effektivitet: {result.efficiency:.0%}")
    print("-" * 100)

    header = f"{'Sol%':<6} {'Vind%':<6} {'Sol CF':<8} {'Vind CF':<8} {'Baseload':<10} {'Baseload%':<10} {'Batt MWh':<10} {'Batt MW':<9} {'Duration':<10} {'OK':<4}"
    print(header)
    print("-" * 100)

    for r in result.all_results:
        sol_pct = f"{r.sol_ratio:.0%}"
        vind_pct = f"{1 - r.sol_ratio:.0%}"
        sol_cf = f"{r.sol_cf:.1%}" if r.sol_mw > 0 else "-"
        vind_cf = f"{r.vind_cf:.1%}" if r.vind_mw > 0 else "-"
        baseload = f"{r.baseload_mw:.3f} MW"
        baseload_pct = f"{r.baseload_pct:.1%}"

        if r.battery_mwh < float('inf'):
            batt_mwh = f"{r.battery_mwh:.2f}"
            batt_mw = f"{r.battery_mw:.3f}"
            duration = f"{r.battery_duration_h:.1f}h"
        else:
            batt_mwh = "inf"
            batt_mw = "-"
            duration = "-"

        ok = "JA" if r.meets_constraint else "NEJ"
        marker = " *" if result.optimal and r.sol_ratio == result.optimal.sol_ratio else ""

        print(f"{sol_pct:<6} {vind_pct:<6} {sol_cf:<8} {vind_cf:<8} {baseload:<10} {baseload_pct:<10} {batt_mwh:<10} {batt_mw:<9} {duration:<10} {ok:<4}{marker}")

    print("-" * 100)

    if result.optimal:
        opt = result.optimal
        print(f"\n* OPTIMAL: {opt.sol_ratio:.0%} Sol / {1 - opt.sol_ratio:.0%} Vind")
        print(f"  Max Baseload: {opt.baseload_mw:.3f} MW ({opt.baseload_pct:.1%} av medelproduktion)")
        print(f"  Batteri: {opt.battery_mwh:.2f} MWh / {opt.battery_mw:.3f} MW ({opt.battery_duration_h:.1f}h duration)")
        print(f"  Kapacitetsfaktor: Sol {opt.sol_cf:.1%}, Vind {opt.vind_cf:.1%}, Kombinerat {opt.combined_cf:.1%}")
    else:
        print("\nIngen lösning hittades inom batteri-constrainten.")

    print("=" * 100)


def generate_heatmap_data(result: RatioOptimizationResult) -> dict:
    """
    Generate data suitable for heatmap visualization.

    Returns dict with:
        - sol_ratios: list of solar ratios (x-axis)
        - baseload_pcts: list of baseload percentages (y-values)
        - battery_durations: list of battery durations
        - meets_constraints: list of booleans
    """
    return {
        "sol_ratios": [r.sol_ratio for r in result.all_results],
        "vind_ratios": [1 - r.sol_ratio for r in result.all_results],
        "baseload_mw": [r.baseload_mw for r in result.all_results],
        "baseload_pcts": [r.baseload_pct for r in result.all_results],
        "battery_mwh": [r.battery_mwh if r.battery_mwh < float('inf') else None for r in result.all_results],
        "battery_durations": [r.battery_duration_h if r.battery_duration_h < float('inf') else None for r in result.all_results],
        "meets_constraints": [r.meets_constraint for r in result.all_results],
        "combined_cf": [r.combined_cf for r in result.all_results],
    }


def export_ratio_optimization_excel(
    result: RatioOptimizationResult,
    filepath: Optional[str] = None,
) -> Path:
    """
    Export optimization results to Excel.

    Creates workbook with:
    - Sheet 1: Summary with optimal configuration
    - Sheet 2: All grid results
    - Sheet 3: Heatmap data for visualization
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if filepath is None:
        filepath = DATA_DIR / "reports" / f"ratio_optimization_{result.zone}_{result.year}.xlsx"
    else:
        filepath = Path(filepath)

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1B4F72')
    highlight_fill = PatternFill('solid', fgColor='E8F6F3')
    input_font = Font(color='0000FF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================================
    # Sheet 1: Summary
    # ========================================
    ws = wb.active
    ws.title = "Sammanfattning"

    ws['A1'] = "OPTIMAL SOL/VIND RATIO FÖR BASELOAD PPA"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # Parameters
    params = [
        ("Zon", result.zone),
        ("År", result.year),
        ("Solprofil", result.sol_profile_name),
        ("Total kapacitet (MW)", result.total_capacity_mw),
        ("Batteri min duration (h)", result.battery_duration_min),
        ("Batteri max duration (h)", result.battery_duration_max),
        ("Batterieffektivitet", f"{result.efficiency:.0%}"),
    ]

    ws['A3'] = "PARAMETRAR"
    ws['A3'].font = Font(bold=True)
    for i, (label, value) in enumerate(params, start=4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value

    # Optimal result
    if result.optimal:
        opt = result.optimal
        ws['A12'] = "OPTIMAL KONFIGURATION"
        ws['A12'].font = Font(bold=True)
        ws['A12'].fill = header_fill
        ws['A12'].font = header_font

        opt_data = [
            ("Sol-andel", f"{opt.sol_ratio:.0%}"),
            ("Vind-andel", f"{1 - opt.sol_ratio:.0%}"),
            ("Sol (MW)", f"{opt.sol_mw:.2f}"),
            ("Vind (MW)", f"{opt.vind_mw:.2f}"),
            ("Max Baseload (MW)", f"{opt.baseload_mw:.3f}"),
            ("Baseload %", f"{opt.baseload_pct:.1%}"),
            ("Batteri (MWh)", f"{opt.battery_mwh:.2f}"),
            ("Batteri (MW)", f"{opt.battery_mw:.3f}"),
            ("Duration (h)", f"{opt.battery_duration_h:.1f}"),
            ("Sol CF", f"{opt.sol_cf:.1%}"),
            ("Vind CF", f"{opt.vind_cf:.1%}"),
            ("Kombinerad CF", f"{opt.combined_cf:.1%}"),
        ]

        for i, (label, value) in enumerate(opt_data, start=13):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].fill = highlight_fill
            ws[f'B{i}'].fill = highlight_fill

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    # ========================================
    # Sheet 2: All Results
    # ========================================
    ws2 = wb.create_sheet("Alla Resultat")

    headers = [
        "Sol%", "Vind%", "Sol MW", "Vind MW", "Sol CF", "Vind CF",
        "Komb CF", "Medelprod MW", "Baseload MW", "Baseload%",
        "Batteri MWh", "Batteri MW", "Duration h", "OK"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, r in enumerate(result.all_results, start=2):
        ws2.cell(row=row_idx, column=1, value=r.sol_ratio).number_format = '0%'
        ws2.cell(row=row_idx, column=2, value=1 - r.sol_ratio).number_format = '0%'
        ws2.cell(row=row_idx, column=3, value=r.sol_mw).number_format = '0.00'
        ws2.cell(row=row_idx, column=4, value=r.vind_mw).number_format = '0.00'
        ws2.cell(row=row_idx, column=5, value=r.sol_cf).number_format = '0.0%'
        ws2.cell(row=row_idx, column=6, value=r.vind_cf).number_format = '0.0%'
        ws2.cell(row=row_idx, column=7, value=r.combined_cf).number_format = '0.0%'
        ws2.cell(row=row_idx, column=8, value=r.mean_production_mw).number_format = '0.000'
        ws2.cell(row=row_idx, column=9, value=r.baseload_mw).number_format = '0.000'
        ws2.cell(row=row_idx, column=10, value=r.baseload_pct).number_format = '0.0%'
        ws2.cell(row=row_idx, column=11, value=r.battery_mwh if r.battery_mwh < float('inf') else None).number_format = '0.00'
        ws2.cell(row=row_idx, column=12, value=r.battery_mw).number_format = '0.000'
        ws2.cell(row=row_idx, column=13, value=r.battery_duration_h if r.battery_duration_h < float('inf') else None).number_format = '0.0'
        ws2.cell(row=row_idx, column=14, value="JA" if r.meets_constraint else "NEJ")

        # Highlight optimal row
        if result.optimal and r.sol_ratio == result.optimal.sol_ratio:
            for col in range(1, 15):
                ws2.cell(row=row_idx, column=col).fill = highlight_fill

    # Freeze header
    ws2.freeze_panes = 'A2'

    # Column widths
    for col in range(1, 15):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 12

    # ========================================
    # Save
    # ========================================
    wb.save(filepath)
    return filepath
